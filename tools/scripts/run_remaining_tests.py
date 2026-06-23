from datetime import date, timedelta, datetime
import os
from pathlib import Path
import sys
import tempfile
import asyncio
import io

# 프로젝트 루트를 경로에 추가
PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

# 테스트용 임시 디렉토리 설정 (기본 DB 보호)
TEST_DATA_DIR = Path(
    os.environ.setdefault("SHIM_DATA_DIR", tempfile.mkdtemp(prefix="shim_full_test_"))
)

from fastapi.testclient import TestClient
from src.app.main import app, startup_event
from src.app.database import SessionLocal
from src.app import models, auth, utils

def next_business_day(start, db):
    d = start
    while True:
        if d.weekday() >= 5:
            d += timedelta(days=1)
            continue
        is_holiday = db.query(models.Holidays).filter(models.Holidays.date == d).first()
        if is_holiday:
            d += timedelta(days=1)
            continue
        break
    return d

def clear_user_leaves(db, user_id):
    db.query(models.Leaves).filter(models.Leaves.user_id == user_id).delete(synchronize_session=False)

def main():
    print(f"[TEST] Starting Full Logic Verification (v1.4.0 Compatible)")
    print(f"[TEST] Database: {TEST_DATA_DIR}")
    
    # 1. 환경 초기화
    from tools.scripts.db_init import init_db
    init_db()
    startup_event()
    client = TestClient(app)
    
    # 초기화 세션 생성 후 즉시 닫기
    db = SessionLocal()

    # 2. 테스트 사용자 생성 (Staff, PM)
    users_to_create = [
        ("u_staff", "일반사원", "STAFF", "Team-A"),
        ("u_pm", "PM관리자", "PM", "HQ"),
        ("u_lead", "팀장", "TEAM_LEAD", "Team-A")
    ]
    for uid, name, role, team in users_to_create:
        if not db.query(models.Users).filter(models.Users.user_id == uid).first():
            db.add(models.Users(
                user_id=uid, user_name=name, role=role, team=team,
                password=auth.get_password_hash("0000"), is_active=True
            ))
    d1 = next_business_day(date.today() + timedelta(days=1), db)
    db.commit()
    db.close() # 유저 생성 후 닫기
    admin_token = auth.create_access_token({"sub": "admin"})
    admin_client = TestClient(app)
    admin_client.cookies.set("access_token", f"Bearer {admin_token}")

    # --- 시나리오 1: PM 자동 승인 (v1.3.x+ 핵심 로직) ---
    print("[CASE 1] PM Auto-Approval Logic")
    # 승인 필수 옵션 ON
    admin_client.post("/api/admin/settings/approval", data={"is_approval_required": "true"})
    
    pm_token = auth.create_access_token({"sub": "u_pm"})
    pm_client = TestClient(app)
    pm_client.cookies.set("access_token", f"Bearer {pm_token}")
    
    # 작업 전 세션 열기/닫기 루틴
    db = SessionLocal()
    clear_user_leaves(db, "u_pm")
    db.commit()
    db.close()

    r = pm_client.post("/api/user/leave", data={
        "date_str": d1.strftime("%Y-%m-%d"), "start_time": "09:00", "end_time": "12:00"
    })
    if r.status_code != 200:
        print(f"  [ERROR] PM Apply failed: {r.status_code} - {r.text}")
    assert r.status_code == 200
    
    db = SessionLocal()
    leave_pm = db.query(models.Leaves).filter(models.Leaves.user_id == "u_pm").first()
    assert leave_pm.status == "APPROVED", f"PM 연차는 자동 승인되어야 합니다. (현재: {leave_pm.status})"
    db.close()
    print("  -> PASS: PM Auto-approval verified.")

    # --- 시나리오 2: 일반 사원 승인 대기 ---
    print("[CASE 2] Staff Pending Logic")
    staff_token = auth.create_access_token({"sub": "u_staff"})
    staff_client = TestClient(app)
    staff_client.cookies.set("access_token", f"Bearer {staff_token}")
    
    db = SessionLocal()
    clear_user_leaves(db, "u_staff")
    db.commit()
    db.close()

    r = staff_client.post("/api/user/leave", data={
        "date_str": d1.strftime("%Y-%m-%d"), "start_time": "14:00", "end_time": "18:00"
    })
    if r.status_code != 200:
        print(f"  [ERROR] Staff Apply failed: {r.status_code} - {r.text}")
    assert r.status_code == 200
    
    db = SessionLocal()
    leave_staff = db.query(models.Leaves).filter(models.Leaves.user_id == "u_staff").first()
    assert leave_staff.status == "PENDING", "일반 사원 연차는 승인 대기 상태여야 합니다."

    # 신규 테스트: 본인 연차 취소 로직 및 타인 연차 취소 시도 검증
    print("  [SUB-CASE] User Leave Cancellation Logic")
    # 취소용 연차 신청
    r_for_cancel = staff_client.post("/api/user/leave", data={
        "date_str": d1.strftime("%Y-%m-%d"), "start_time": "09:00", "end_time": "12:00"
    })
    assert r_for_cancel.status_code == 200
    
    db = SessionLocal()
    leave_to_cancel = db.query(models.Leaves).filter(models.Leaves.user_id == "u_staff", models.Leaves.snapshot_start_min == 540).first()
    assert leave_to_cancel is not None
    leave_to_cancel_id = leave_to_cancel.id
    db.close()
    
    # 1) 타인(PM)이 u_staff의 연차를 취소 시도 -> 403
    r_cancel_by_other = pm_client.post(f"/api/user/leave/cancel/{leave_to_cancel_id}")
    assert r_cancel_by_other.status_code == 403
    
    # 2) 본인(u_staff)이 본인의 연차를 취소 -> 200
    r_cancel_by_self = staff_client.post(f"/api/user/leave/cancel/{leave_to_cancel_id}")
    assert r_cancel_by_self.status_code == 200
    
    db = SessionLocal()
    canceled_leave = db.query(models.Leaves).filter(models.Leaves.id == leave_to_cancel_id).first()
    assert canceled_leave.status == "CANCELED"
    db.close()
    print("  -> PASS: User cancellation logic verified.")
    
    # --- 시나리오 3: 연차 삭제 시 감사 로그 생성 및 Race Condition 방지 (v1.4.0) ---
    print("[CASE 3] Leave Deletion & Audit Log (v1.4.0)")
    leave_id = leave_staff.id

    # [검증] 공가 전환 후 연차 복구 전환 시 에러 여부 검증
    # 1) 공가(비차감)로 전환
    r_to_public = admin_client.post("/api/admin/leave/update-type", data={"leave_id": leave_id, "is_deductive": "false"})
    assert r_to_public.status_code == 200
    db = SessionLocal()
    leave_checked = db.query(models.Leaves).filter(models.Leaves.id == leave_id).first()
    assert leave_checked.is_deductive is False, "공가 전환 실패"
    db.close()

    # 2) 다시 연차(차감)로 복구 전환
    r_to_deduct = admin_client.post("/api/admin/leave/update-type", data={"leave_id": leave_id, "is_deductive": "true"})
    assert r_to_deduct.status_code == 200
    db = SessionLocal()
    leave_checked = db.query(models.Leaves).filter(models.Leaves.id == leave_id).first()
    assert leave_checked.is_deductive is True, "연차 전환 실패"

    # 3) 감사 로그 기록 검증
    audit_type = db.query(models.AuditLogs).filter(models.AuditLogs.action == "UPDATE_LEAVE_TYPE").order_by(models.AuditLogs.id.desc()).first()
    assert audit_type is not None, "휴가 유형 변경 감사 로그 누락"
    assert "is_deductive=True" in audit_type.new_data
    db.close()

    # Admin으로 연차 삭제 수행
    r = admin_client.post("/api/admin/leave/delete", data={"leave_id": leave_id})
    assert r.status_code in (200, 302)
    
    db = SessionLocal()
    # 삭제 확인
    deleted = db.query(models.Leaves).filter(models.Leaves.id == leave_id).first()
    assert deleted is None, "연차가 삭제되지 않았습니다."
    
    # 감사 로그 확인 (v1.4.0에서 액션명이 LEAVE_DELETE 또는 DELETE_LEAVE일 수 있음, 확인 결과 api_admin.py는 LEAVE_DELETE 사용)
    audit = db.query(models.AuditLogs).filter(models.AuditLogs.action == "LEAVE_DELETE").order_by(models.AuditLogs.id.desc()).first()
    if not audit: # 폴백 체크
         audit = db.query(models.AuditLogs).filter(models.AuditLogs.action == "DELETE_LEAVE").order_by(models.AuditLogs.id.desc()).first()
    
    assert audit is not None, "삭제 감사 로그가 생성되지 않았습니다."
    print("  -> PASS: Deletion & Audit log verified.")

    # --- 시나리오 4: 1.4.0 UI 전역 함수 (min/max) 및 페이징 쿼리 검증 ---
    print("[CASE 4] Pagination & Template Globals Check")
    # 타임라인 쿼리 (admin_service)
    from src.app.services import admin_service
    query = admin_service.get_leaves_timeline_query(db, year=d1.year)
    # 페이징 적용 확인 (v1.4.0에서 50건 단위 도입됨)
    paged_results = query.offset(0).limit(50).all()
    assert isinstance(paged_results, list)
    
    # Template globals check (min/max)
    assert app.state.templates.env.globals["min"] == min
    assert app.state.templates.env.globals["max"] == max
    print("  -> PASS: v1.4.0 Architecture & Logic verified.")

    # --- 시나리오 5: 전사 캘린더 공유 활성화 및 권한 검증 ---
    print("[CASE 5] Company Calendar Sharing & Access Control")
    
    # 1. 캘린더 공유 범위를 '소속 팀원 공유 (team)'로 변경
    admin_client.post("/api/admin/settings/calendar-scope", data={"scope": "team"})
    
    from src.app.routers.api_user import user_team_calendar
    from unittest.mock import MagicMock
    
    # 가짜 Request 객체 생성
    req = MagicMock()
    req.url.path = "/user/team-calendar"
    req.app.state.templates = app.state.templates
    
    from src.app.routers import api_user as api_user_module
    orig_get_current_user = api_user_module.get_current_user
    
    db = SessionLocal()
    staff_user = db.query(models.Users).filter(models.Users.user_id == "u_staff").first()
    db.close()
    
    # 1-1. scope=team 일 때 (본인 팀원만 조회되어야 함)
    api_user_module.get_current_user = lambda r, d: staff_user
    db = SessionLocal()
    resp = user_team_calendar(request=req, db=db, user=staff_user)
    context = resp.context
    members = context["team_members"]
    member_ids = [m.user_id for m in members]
    assert "u_staff" in member_ids
    assert "u_lead" in member_ids
    assert "u_pm" not in member_ids, f"팀원 캘린더 공유 상태에서 다른 팀원(PM)이 조회되면 안 됩니다. (조회 목록: {member_ids})"
    db.close()
    
    # 2. 캘린더 공유 범위를 '전사 공유 (company)'로 변경
    r = admin_client.post("/api/admin/settings/calendar-scope", data={"scope": "company"})
    assert r.status_code == 200
    
    # 2-1. 감사 로그 생성 검증
    db = SessionLocal()
    audit_setting = db.query(models.AuditLogs).filter(models.AuditLogs.action == "UPDATE_CALENDAR_SCOPE_SETTING").order_by(models.AuditLogs.id.desc()).first()
    assert audit_setting is not None, "캘린더 공유 범위 변경 감사 로그가 생성되지 않았습니다."
    assert audit_setting.new_data == "company"
    db.close()
    
    # 2-2. scope=company 일 때 (전사 사원 모두 조회되어야 함)
    db = SessionLocal()
    resp = user_team_calendar(request=req, db=db, user=staff_user)
    context = resp.context
    members = context["team_members"]
    member_ids = [m.user_id for m in members]
    assert "u_staff" in member_ids
    assert "u_lead" in member_ids
    assert "u_pm" in member_ids, f"전사 캘린더 공유 활성화 시 다른 팀원(PM)도 조회되어야 합니다. (조회 목록: {member_ids})"
    db.close()
 
    # 3. 캘린더 공유 범위를 '공유 안 함 (none)'으로 변경
    r = admin_client.post("/api/admin/settings/calendar-scope", data={"scope": "none"})
    assert r.status_code == 200
    
    # 3-1. 리다이렉트 응답 확인
    db = SessionLocal()
    resp = user_team_calendar(request=req, db=db, user=staff_user)
    assert resp.status_code == 302
    assert resp.headers.get("location") == "/user/dashboard"
    db.close()
    
    # 원래대로 복구
    api_user_module.get_current_user = orig_get_current_user
    print("  -> PASS: Company Calendar Sharing & Access Control verified.")

    # --- 시나리오 6: 다수일 일괄 신청 및 롤백 검증 (v1.5.0) ---
    print("[CASE 6] Bulk Leave Application & Rollback (v1.5.0)")
    
    # 점심시간 설정 변경 (12:00 ~ 13:00)
    admin_client.post("/api/admin/settings/time-policy", data={
        "time_granularity_minutes": 60,
        "work_start_minute": 540,
        "work_end_minute": 1080,
        "lunch_start_minute": 720,
        "lunch_end_minute": 780
    })
    
    # u_staff의 기존 연차 삭제
    db = SessionLocal()
    clear_user_leaves(db, "u_staff")
    db.commit()
    db.close()
    
    # 2026-06-08(월) ~ 2026-06-14(일) 중 평일: 5일 (08, 09, 10, 11, 12). 주말: 13, 14
    # 콤마로 구분된 날짜 리스트 (주말인 13, 14를 포함시킴)
    bulk_dates = "2026-06-08,2026-06-09,2026-06-10,2026-06-11,2026-06-12,2026-06-13,2026-06-14"
    
    r = staff_client.post("/api/user/leave", data={
        "date_str": bulk_dates, "start_time": "09:00", "end_time": "18:00"
    })
    
    assert r.status_code == 200, f"일괄 신청 실패: {r.text}"
    
    # DB 확인: 6월 13일(토), 14일(일)은 스킵되고 평일 5개만 등록되어야 함.
    db = SessionLocal()
    leaves = db.query(models.Leaves).filter(models.Leaves.user_id == "u_staff").all()
    assert len(leaves) == 5, f"평일 5일만 등록되어야 하는데 {len(leaves)}개가 등록되었습니다."
    
    # 감사 로그 확인
    audit_bulk = db.query(models.AuditLogs).filter(models.AuditLogs.action == "APPLY_LEAVE_BULK").order_by(models.AuditLogs.id.desc()).first()
    assert audit_bulk is not None, "일괄 신청 감사 로그가 없습니다."
    assert "Leaves for 2026-06-" in audit_bulk.target_info, "감사 로그 타겟 정보 확인 실패"
    
    # 롤백 검증: 기존 등록된 6월 8일이 중복되도록 새로운 일괄 신청
    # 2026-06-08(중복), 2026-06-15(미등록) 신청
    r_dup = staff_client.post("/api/user/leave", data={
        "date_str": "2026-06-08,2026-06-15", "start_time": "09:00", "end_time": "18:00"
    })
    
    assert r_dup.status_code == 400, "중복 신청이 포함되었는데 에러가 발생하지 않았습니다."
    
    # 2026-06-15가 등록되지 않았는지 확인 (롤백 여부)
    leave_15 = db.query(models.Leaves).filter(
        models.Leaves.user_id == "u_staff", 
        models.Leaves.date == date(2026, 6, 15)
    ).first()
    assert leave_15 is None, "롤백이 수행되지 않아 2026-06-15 연차가 등록되었습니다."
    
    # --- v1.5.1 추가 검증: 시간 생략 전송 (자동 하루종일 매핑) ---
    clear_user_leaves(db, "u_staff")
    db.commit()
    
    # 1) 다수일 일괄 신청 시 시간 생략
    r_all_day_bulk = staff_client.post("/api/user/leave", data={
        "date_str": "2026-06-08,2026-06-09",
        "start_time": "",
        "end_time": ""
    })
    assert r_all_day_bulk.status_code == 200, f"시간 생략 다수일 신청 실패: {r_all_day_bulk.text}"
    
    leaves_bulk = db.query(models.Leaves).filter(
        models.Leaves.user_id == "u_staff",
        models.Leaves.date.in_([date(2026, 6, 8), date(2026, 6, 9)])
    ).all()
    assert len(leaves_bulk) == 2
    for lv in leaves_bulk:
        assert lv.snapshot_deduction_hours == 8.0, f"예상 차감 시간: 8.0, 실제: {lv.snapshot_deduction_hours}"
        assert lv.snapshot_start_min == 540, f"예상 시작 분: 540, 실제: {lv.snapshot_start_min}"
        assert lv.snapshot_end_min == 1080, f"예상 종료 분: 1080, 실제: {lv.snapshot_end_min}"
        
    clear_user_leaves(db, "u_staff")
    db.commit()
    
    # 2) 단일 날짜 신청 시 시간 생략 (하루종일 체크박스 전송 시나리오)
    r_all_day_single = staff_client.post("/api/user/leave", data={
        "date_str": "2026-06-08",
        "start_time": "",
        "end_time": ""
    })
    assert r_all_day_single.status_code == 200, f"시간 생략 단일 신청 실패: {r_all_day_single.text}"
    
    leave_single = db.query(models.Leaves).filter(
        models.Leaves.user_id == "u_staff",
        models.Leaves.date == date(2026, 6, 8)
    ).first()
    assert leave_single is not None
    assert leave_single.snapshot_deduction_hours == 8.0
    assert leave_single.snapshot_start_min == 540
    assert leave_single.snapshot_end_min == 1080
    
    db.close()
    print("  -> PASS: Bulk leave application & Rollback verified (including v1.5.1 time-omitted Full-day mapping).")
    
    # --- 시나리오 7: 시스템 설정 인메모리 캐싱 검증 (v1.5.0) ---
    print("[CASE 7] System Settings In-Memory Caching (v1.5.0)")
    
    from src.app.services.leave_policy import get_system_settings
    
    db = SessionLocal()
    # 초기 캐싱 동작 확인
    settings_before = get_system_settings(db)
    initial_title = settings_before.product_display_name
    
    # Admin API를 통한 설정 변경 (Branding 명칭 변경)
    import time
    new_title = f"SHIM-{int(time.time())}"
    # app.css 등을 로드하는 과정에서 캐시 설정이 변경되므로 설정 API에 요청 보냄
    r_setting = admin_client.post("/api/admin/settings/branding", data={
        "product_display_name": new_title,
        "product_nav_short": settings_before.product_nav_short or "",
        "brand_initial": settings_before.brand_initial or ""
    })
    assert r_setting.status_code == 200, "설정 변경 API 호출 실패"
    
    # 캐시가 갱신되어 새로운 타이틀이 나오는지 확인
    settings_after = get_system_settings(db)
    assert settings_after.product_display_name == new_title, f"인메모리 캐시 갱신 실패. (예상: {new_title}, 실제: {settings_after.product_display_name})"
    db.close()
    print("  -> PASS: System settings cache verified.")

    # --- 시나리오 8: 비밀번호 변경 기능 검증 (자가 변경 및 관리자 변경) ---
    print("[CASE 8] User & Admin Password Self-Service Change")
    
    # 1) 일반 사용자 비밀번호 변경 검증
    # 잘못된 현재 비밀번호 입력 시 실패
    r = staff_client.post("/api/user/change-password", data={
        "current_password": "wrong_password",
        "new_password": "new_password_123"
    })
    assert r.status_code == 400
    assert "현재 비밀번호가 일치하지 않습니다." in r.json()["message"]
    
    # 새 비밀번호 길이 미달(8자 미만) 시 실패
    r = staff_client.post("/api/user/change-password", data={
        "current_password": "0000",
        "new_password": "abc"
    })
    assert r.status_code == 400
    assert "최소 8자 이상" in r.json()["message"]

    # 성공적인 비밀번호 변경
    r = staff_client.post("/api/user/change-password", data={
        "current_password": "0000",
        "new_password": "NewStaffPass123!"
    })
    assert r.status_code == 200
    assert "비밀번호가 성공적으로 변경되었습니다." in r.json()["message"]
    
    # 변경된 비밀번호로 검증
    db = SessionLocal()
    updated_user = db.query(models.Users).filter(models.Users.user_id == "u_staff").first()
    assert auth.verify_password("NewStaffPass123!", updated_user.password)
    
    # 감사 로그(CHANGE_PASSWORD) 확인
    audit_pwd = db.query(models.AuditLogs).filter(
        models.AuditLogs.action == "CHANGE_PASSWORD",
        models.AuditLogs.actor_id == "u_staff"
    ).order_by(models.AuditLogs.id.desc()).first()
    assert audit_pwd is not None
    assert audit_pwd.old_data == "*****"
    assert audit_pwd.new_data == "*****"
    db.close()

    # 2) 관리자 비밀번호 변경 검증
    # 잘못된 현재 비밀번호
    r = admin_client.post("/api/admin/change-password", data={
        "current_password": "wrong_admin_pwd",
        "new_password": "new_admin_pwd_123"
    })
    assert r.status_code == 400
    assert "현재 비밀번호가 일치하지 않습니다." in r.json()["message"]
    
    # 새 비밀번호 길이 미달(8자 미만)
    r = admin_client.post("/api/admin/change-password", data={
        "current_password": "0000",
        "new_password": "123"
    })
    assert r.status_code == 400
    assert "최소 8자 이상" in r.json()["message"]

    # 성공적인 변경
    r = admin_client.post("/api/admin/change-password", data={
        "current_password": "0000",
        "new_password": "NewAdminPass123!"
    })
    assert r.status_code == 200
    
    db = SessionLocal()
    updated_admin = db.query(models.Users).filter(models.Users.user_id == "admin").first()
    assert auth.verify_password("NewAdminPass123!", updated_admin.password)
    
    # 감사 로그(CHANGE_ADMIN_PASSWORD) 확인
    audit_admin = db.query(models.AuditLogs).filter(
        models.AuditLogs.action == "CHANGE_ADMIN_PASSWORD"
    ).order_by(models.AuditLogs.id.desc()).first()
    assert audit_admin is not None
    assert audit_admin.old_data == "*****"
    assert audit_admin.new_data == "*****"
    db = SessionLocal()
    for uid in ("u_staff", "admin"):
        usr = db.query(models.Users).filter(models.Users.user_id == uid).first()
        if usr:
            usr.token_version = 0
    db.commit()
    db.close()
    
    print("  -> PASS: User & Admin password change verified.")

    # --- 시나리오 9: SQLite 외래 키 제약 조건 및 트랜잭션 롤백 무결성 검증 ---
    print("[CASE 9] SQLite Foreign Key & Rollback Integrity")
    db = SessionLocal()
    # 외래 키 위반 시 유발 검증: 존재하지 않는 user_id인 'non_existent_user'로 Leaves 삽입 시도
    invalid_leave = models.Leaves(
        user_id="non_existent_user", # 존재하지 않는 사용자 ID
        date=date.today(),
        snapshot_slot_label="09:00~18:00",
        snapshot_start_min=540,
        snapshot_end_min=1080,
        snapshot_deduction_hours=8.0,
        status="APPROVED",
        year=date.today().year,
        is_deductive=True
    )
    db.add(invalid_leave)
    from sqlalchemy.exc import IntegrityError
    try:
        db.commit()
        # 실패해야 정상
        assert False, "외래 키 제약 조건 위반 에러가 발생해야 합니다."
    except IntegrityError:
        db.rollback()
        print("  -> PASS: Foreign key validation error and rollback successfully handled.")
    db.close()

    # --- 시나리오 10: 소수점 연차 잔여량 표기 및 유틸리티 정밀성 검증 ---
    print("[CASE 10] Decimal Precision Formatting (utils)")
    from src.app import utils as shim_utils
    # 0.5단위 소수점이 온전히 출력되는지 검증 (반올림으로 뭉개지지 않음)
    assert shim_utils.hours_to_days_hours_compact(7.5) == "7.5h"
    assert shim_utils.hours_to_days_hours_compact(-4.5) == "-4.5h"
    assert shim_utils.hours_to_days_hours_compact(8.0) == "1일"
    assert shim_utils.hours_to_days_hours_compact(12.0) == "1일4h"
    assert shim_utils.hours_to_days_hours_compact(12.5) == "1일4.5h"
    print("  -> PASS: Decimal formatting precision verified.")

    # --- 시나리오 11: JWT 쿠키 보안 설정 및 HTTPS 프로토콜 분기 검증 ---
    print("[CASE 11] JWT Cookie Security Settings")
    # 1. 일반 로그인 응답에서 SameSite=Lax 확인
    cookie_client = TestClient(app)
    r_login = cookie_client.post("/login", data={"user_id": "u_staff", "password": "NewStaffPass123!"}, follow_redirects=False)
    assert r_login.status_code == 302
    cookie_header = r_login.headers.get("set-cookie", "")
    assert "samesite=lax" in cookie_header.lower()
    
    # 2. SHIM_SECURE_COOKIE 활성화 시 Secure=True 주입 검증
    os.environ["SHIM_SECURE_COOKIE"] = "true"
    secure_cookie_client = TestClient(app)
    r_secure_login = secure_cookie_client.post("/login", data={"user_id": "u_staff", "password": "NewStaffPass123!"}, follow_redirects=False)
    cookie_header_sec = r_secure_login.headers.get("set-cookie", "")
    assert "secure" in cookie_header_sec.lower()
    os.environ.pop("SHIM_SECURE_COOKIE", None)
    print("  -> PASS: SameSite=Lax and Secure cookie dynamic flag verified.")

    # --- 시나리오 12: SQLite WAL 백업 기능 동작성 검증 ---
    print("[CASE 12] SQLite WAL Online Backup Operation")
    from src.app.services import ops
    from src.app.database import DB_PATH
    
    backup_dir = TEST_DATA_DIR / "backups"
    
    backup_path = ops.create_sqlite_backup(db_path=DB_PATH, backup_dir=backup_dir)
    assert backup_path.exists(), "백업 파일이 생성되지 않았습니다."
    assert backup_path.stat().st_size > 0, "백업 파일 크기가 0 바이트입니다."
    backup_path.unlink()
    print("  -> PASS: Online backup using sqlite3.backup() completed successfully.")

    # --- 시나리오 13: 역할 기반 권한 제어(RBAC) 및 보안 정책 우회 차단 검증 ---
    print("[CASE 13] Role-Based Access Control (RBAC) & Security Bypassing")
    # 1) 일반 사원(STAFF) 권한으로 어드민 전용 URL 접속 차단 검증
    staff_client_req = TestClient(app)
    staff_client_req.cookies.set("access_token", f"Bearer {staff_token}")
    r_admin_dash = staff_client_req.get("/admin/dashboard", follow_redirects=False)
    assert r_admin_dash.status_code in (302, 403)
    
    # 2) 팀장(TEAM_LEAD)이 자신의 연차 신청 건에 대해 결재(Self-approval)를 시도하는 시나리오 차단 검증
    lead_token = auth.create_access_token({"sub": "u_lead"})
    lead_client = TestClient(app)
    lead_client.cookies.set("access_token", f"Bearer {lead_token}")
    
    db = SessionLocal()
    clear_user_leaves(db, "u_lead")
    db.commit()
    db.close()
    
    r_lead_apply = lead_client.post("/api/user/leave", data={
        "date_str": d1.strftime("%Y-%m-%d"), "start_time": "09:00", "end_time": "18:00"
    })
    assert r_lead_apply.status_code == 200
    
    db = SessionLocal()
    lead_leave = db.query(models.Leaves).filter(models.Leaves.user_id == "u_lead").first()
    assert lead_leave is not None
    assert lead_leave.status == "PENDING"
    db.close()
    
    r_self_approve = lead_client.post(f"/api/user/team-approve/{lead_leave.id}")
    assert r_self_approve.status_code in (400, 403)
    
    # 3) 팀장(TEAM_LEAD)이 다른 팀 소속 사원의 연차에 대해 결재를 시도하는 시나리오 차단 검증
    db = SessionLocal()
    lead_b = db.query(models.Users).filter(models.Users.user_id == "u_lead_b").first()
    if not lead_b:
        db.add(models.Users(
            user_id="u_lead_b", user_name="팀장B", role="TEAM_LEAD", team="Team-B",
            password=auth.get_password_hash("0000"), is_active=True
        ))
        db.commit()
    db.close()
    
    db = SessionLocal()
    clear_user_leaves(db, "u_staff")
    db.commit()
    db.close()
    
    r_staff_apply = staff_client.post("/api/user/leave", data={
        "date_str": d1.strftime("%Y-%m-%d"), "start_time": "09:00", "end_time": "18:00"
    })
    assert r_staff_apply.status_code == 200
    
    db = SessionLocal()
    staff_leave = db.query(models.Leaves).filter(models.Leaves.user_id == "u_staff").first()
    db.close()
    
    lead_b_token = auth.create_access_token({"sub": "u_lead_b"})
    lead_b_client = TestClient(app)
    lead_b_client.cookies.set("access_token", f"Bearer {lead_b_token}")
    
    r_other_approve = lead_b_client.post(f"/api/user/team-approve/{staff_leave.id}")
    assert r_other_approve.status_code == 403
    print("  -> PASS: RBAC check (Admin page block, self-approval block, other team block) verified.")

    # --- 시나리오 14: 정밀 시간 차감 정책 및 점심시간 제외 바운더리 검증 ---
    print("[CASE 14] Precise Time Policy Boundaries")
    # 1) 60분 단위 경계에 어긋나는 신청(예: 09:30 신청) 시 실패 검증
    r_invalid_time = staff_client.post("/api/user/leave", data={
        "date_str": d1.strftime("%Y-%m-%d"), "start_time": "09:30", "end_time": "11:30"
    })
    assert r_invalid_time.status_code == 400
    assert "시간 단위 경계" in r_invalid_time.json()["message"]
    
    # 2) 점심시간(12:00~13:00)을 걸쳐 연차를 신청한 경우(예: 11:00~14:00) 1시간 제외하고 2.0시간만 차감 검증
    db = SessionLocal()
    clear_user_leaves(db, "u_staff")
    db.commit()
    db.close()
    
    r_lunch_overlap = staff_client.post("/api/user/leave", data={
        "date_str": d1.strftime("%Y-%m-%d"), "start_time": "11:00", "end_time": "14:00"
    })
    assert r_lunch_overlap.status_code == 200
    
    db = SessionLocal()
    lunch_leave = db.query(models.Leaves).filter(models.Leaves.user_id == "u_staff").first()
    assert lunch_leave is not None
    assert lunch_leave.snapshot_deduction_hours == 2.0
    db.close()

    # 3) 점심시간과 완전히 겹치는 신청(예: 12:00~13:00) 시 상세한 에러 안내 및 400 검증
    r_fully_lunch = staff_client.post("/api/user/leave", data={
        "date_str": d1.strftime("%Y-%m-%d"), "start_time": "12:00", "end_time": "13:00"
    })
    assert r_fully_lunch.status_code == 400
    assert "점심시간" in r_fully_lunch.json()["message"]
    assert "완전히 포함" in r_fully_lunch.json()["message"]

    print("  -> PASS: Granularity boundaries and lunch exclusion verified.")

    # --- 시나리오 15: 퇴사자(비활성 사원) 로그인 원천 차단 검증 ---
    print("[CASE 15] Deactivated Employee Login Blocking")
    db = SessionLocal()
    staff_user_obj = db.query(models.Users).filter(models.Users.user_id == "u_staff").first()
    staff_user_obj.is_active = False
    db.commit()
    db.close()
    
    deactive_client = TestClient(app)
    r_deactive_login = deactive_client.post("/login", data={"user_id": "u_staff", "password": "NewStaffPass123!"})
    assert r_deactive_login.status_code == 200
    assert "비활성" in r_deactive_login.text
    
    db = SessionLocal()
    staff_user_obj = db.query(models.Users).filter(models.Users.user_id == "u_staff").first()
    staff_user_obj.is_active = True
    db.commit()
    db.close()
    print("  -> PASS: Deactivated employee login successfully blocked.")

    # --- 시나리오 16: 수동 공휴일 지정에 따른 연차 신청 제한 검증 ---
    print("[CASE 16] Holiday Date Leave Application Blocking")
    test_holiday_date = date(2026, 10, 14)
    db = SessionLocal()
    db.query(models.Holidays).filter(models.Holidays.date == test_holiday_date).delete()
    db.commit()
    db.close()
    
    r_add_holiday = admin_client.post("/api/admin/holiday/create", data={
        "holiday_name": "테스트창립기념일",
        "holiday_date": "2026-10-14"
    })
    assert r_add_holiday.status_code in (200, 302)
    
    r_holiday_apply = staff_client.post("/api/user/leave", data={
        "date_str": "2026-10-14", "start_time": "09:00", "end_time": "18:00"
    })
    assert r_holiday_apply.status_code == 400
    assert "\uacf5\ud734\uc77c" in r_holiday_apply.json()["message"]
    
    db = SessionLocal()
    db.query(models.Holidays).filter(models.Holidays.date == test_holiday_date).delete()
    db.commit()
    db.close()
    print("  -> PASS: Leave application on registered holidays successfully blocked.")

    # --- 시나리오 17: 사원 강제 삭제 시 감사 로그 actor_id NULL 처리 및 정합성 검증 ---
    print("[CASE 17] User Hard-Delete with Audit Logs Integrity")
    db = SessionLocal()
    test_uid = "u_temp_delete"
    
    # 1. 이전 잔재가 있다면 청소
    db.query(models.AuditLogs).filter(models.AuditLogs.actor_id == test_uid).delete()
    db.query(models.Leaves).filter(models.Leaves.user_id == test_uid).delete()
    db.query(models.UserYearlyLeaveAllocations).filter(models.UserYearlyLeaveAllocations.user_id == test_uid).delete()
    db.query(models.Users).filter(models.Users.user_id == test_uid).delete()
    db.commit()
    
    # 2. 신규 사용자 및 연관 데이터 생성
    db.add(models.Users(
        user_id=test_uid, user_name="임시사원", role="STAFF",
        company="본사", team="개발팀",
        password=auth.get_password_hash("0000"), is_active=True
    ))
    db.commit()
    
    # 3. 감사 로그 생성 (이 사용자가 actor가 됨)
    db.add(models.AuditLogs(
        actor_id=test_uid, action="TEST_ACTION_BY_USER",
        target_info="Test Target", old_data="old", new_data="new"
    ))
    # 4. 연차 할당 및 신청 데이터 생성
    db.add(models.UserYearlyLeaveAllocations(user_id=test_uid, year=2026, allocated_hours=120))
    db.add(models.Leaves(
        user_id=test_uid, date=date(2026, 11, 20), snapshot_slot_label="09:00~18:00",
        snapshot_start_min=540, snapshot_end_min=1080, snapshot_deduction_hours=8.0,
        status="APPROVED", year=2026
    ))
    db.commit()
    
    # 생성된 감사 로그 ID 확보
    audit_row = db.query(models.AuditLogs).filter(models.AuditLogs.actor_id == test_uid).first()
    assert audit_row is not None
    audit_id = audit_row.id
    db.close()
    
    # 5. 관리자 권한으로 사원 강제 삭제 API 호출
    r_delete = admin_client.post("/api/admin/user/hard-delete", data={"target_user_id": test_uid})
    assert r_delete.status_code == 200
    
    # 6. DB 정합성 최종 검증
    db = SessionLocal()
    # 사원 레코드 삭제 확인
    assert db.query(models.Users).filter(models.Users.user_id == test_uid).first() is None
    # 연차 및 할당 삭제 확인
    assert db.query(models.Leaves).filter(models.Leaves.user_id == test_uid).all() == []
    assert db.query(models.UserYearlyLeaveAllocations).filter(models.UserYearlyLeaveAllocations.user_id == test_uid).all() == []
    # 감사 로그는 남아있되, actor_id만 NULL로 정상 업데이트 되었는지 검증
    audit_after = db.query(models.AuditLogs).filter(models.AuditLogs.id == audit_id).first()
    assert audit_after is not None
    assert audit_after.actor_id is None
    assert audit_after.actor_name == "임시사원", f"수행자 이름 스냅샷 유실: {audit_after.actor_name}"
    assert audit_after.actor_department == "본사 개발팀", f"수행자 부서 스냅샷 유실: {audit_after.actor_department}"
    
    # 청소
    db.query(models.AuditLogs).filter(models.AuditLogs.id == audit_id).delete()
    db.commit()
    db.close()
    print("  -> PASS: User hard-delete with audit logs integrity verified successfully.")

    # --- 시나리오 18: 사원 권한/상태 변경 시 세션 강제 만료 검증 ---
    print("[CASE 18] Session Invalidation on Role/Status Change")
    
    db = SessionLocal()
    staff_db_user = db.query(models.Users).filter(models.Users.user_id == "u_staff").first()
    assert staff_db_user is not None
    staff_db_user.token_version = 0
    staff_db_user.role = "STAFF"
    staff_db_user.is_active = True
    db.commit()
    db.close()
    
    token_v0 = auth.create_access_token({"sub": "u_staff", "token_version": 0})
    staff_client_v0 = TestClient(app)
    staff_client_v0.cookies.set("access_token", f"Bearer {token_v0}")
    
    r_dash = staff_client_v0.get("/user/dashboard")
    assert r_dash.status_code == 200
    
    r_update = admin_client.post("/api/admin/user/update", data={
        "target_user_id": "u_staff",
        "user_name": "일반사원수정",
        "company": "Team-A",
        "team": "Team-A",
        "role": "TEAM_LEAD",
        "position": "팀장",
        "is_active": "true"
    })
    assert r_update.status_code == 200
    
    r_dash_after = staff_client_v0.get("/user/dashboard", follow_redirects=False)
    assert r_dash_after.status_code in (302, 401)
    
    db = SessionLocal()
    staff_db_user = db.query(models.Users).filter(models.Users.user_id == "u_staff").first()
    curr_version = staff_db_user.token_version
    db.close()
    
    token_v_curr = auth.create_access_token({"sub": "u_staff", "token_version": curr_version})
    staff_client_curr = TestClient(app)
    staff_client_curr.cookies.set("access_token", f"Bearer {token_v_curr}")
    
    r_dash_curr = staff_client_curr.get("/user/dashboard")
    assert r_dash_curr.status_code == 200
    
    r_toggle = admin_client.post("/api/admin/user/toggle", data={"target_user_id": "u_staff"})
    assert r_toggle.status_code == 200
    
    r_dash_curr_after = staff_client_curr.get("/user/dashboard", follow_redirects=False)
    assert r_dash_curr_after.status_code in (302, 401)
    
    db = SessionLocal()
    staff_db_user = db.query(models.Users).filter(models.Users.user_id == "u_staff").first()
    staff_db_user.is_active = True
    staff_db_user.role = "STAFF"
    staff_db_user.token_version = 0
    db.commit()
    db.close()
    print("  -> PASS: Session Invalidation on Role/Status Change verified.")

    # --- 시나리오 19: 엑셀 타임라인 파일 타입 포맷팅 정밀화 검증 ---
    print("[CASE 19] Excel Export Data Type Formatting Refinement")
    
    r_export = admin_client.get(f"/admin/leave/timeline/export?year={d1.year}")
    assert r_export.status_code == 200
    assert r_export.headers.get("content-type") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    from openpyxl import load_workbook
    
    wb = load_workbook(io.BytesIO(r_export.content))
    
    ws_sum = wb["연차 현황 요약"]
    assert isinstance(ws_sum.cell(row=2, column=8).value, (int, float))
    assert isinstance(ws_sum.cell(row=2, column=9).value, (int, float))
    assert isinstance(ws_sum.cell(row=2, column=10).value, (int, float))
    assert isinstance(ws_sum.cell(row=2, column=11).value, (int, float))
    
    ws_time = wb["상세 신청 내역"]
    if ws_time.max_row > 1:
        created_val = ws_time.cell(row=2, column=1).value
        date_val = ws_time.cell(row=2, column=6).value
        deduct_val = ws_time.cell(row=2, column=9).value
        
        assert isinstance(created_val, (datetime, date))
        assert isinstance(date_val, (datetime, date))
        assert isinstance(deduct_val, (int, float))
        
        assert ws_time.cell(row=2, column=1).number_format == 'yyyy-mm-dd hh:mm:ss'
        assert ws_time.cell(row=2, column=6).number_format == 'yyyy-mm-dd'

    print("  -> PASS: Excel data type and formatting verified successfully.")

    # --- 시나리오 20: 사원 관리 목록 권한별 정렬 및 비활성 사원 최하단 고정 검증 (v1.5.17) ---
    print("[CASE 20] Admin Users Sorting & Inactive Exclusion (v1.5.17)")
    
    # 1. 테스트 유저 생성
    db = SessionLocal()
    
    # 기존 테스트 유저들 충돌 방지를 위해 임시 유저 삭제
    temp_uids = ["t_admin", "t_pm", "t_lead", "t_staff1", "t_staff2", "t_inactive1", "t_inactive2"]
    db.query(models.Users).filter(models.Users.user_id.in_(temp_uids)).delete(synchronize_session=False)
    db.commit()
    
    # 테스트 유저 추가
    test_users = [
        models.Users(user_id="t_admin", user_name="가_어드민", role="ADMIN", is_active=True, password="hash"),
        models.Users(user_id="t_pm", user_name="나_피엠", role="PM", is_active=True, password="hash"),
        models.Users(user_id="t_lead", user_name="다_팀장", role="TEAM_LEAD", is_active=True, password="hash"),
        models.Users(user_id="t_staff2", user_name="마_사원", role="STAFF", is_active=True, password="hash"),
        models.Users(user_id="t_staff1", user_name="라_사원", role="STAFF", is_active=True, password="hash"),
        models.Users(user_id="t_inactive1", user_name="바_비활성_스태프", role="STAFF", is_active=False, password="hash"),
        models.Users(user_id="t_inactive2", user_name="사_비활성_어드민", role="ADMIN", is_active=False, password="hash"),
    ]
    for tu in test_users:
        db.add(tu)
    db.commit()
    
    # admin user (get_current_admin Depends에 쓰일 관리자 계정) 확보
    admin_user = db.query(models.Users).filter(models.Users.role == "ADMIN", models.Users.is_active == True).first()
    db.close()
    
    from src.app.routers.admin.users import admin_users
    from unittest.mock import MagicMock
    
    req_admin = MagicMock()
    req_admin.app.state.templates = app.state.templates
    
    # 2. 기본 정렬 검증 (기본 sort_key="role", sort_dir="asc")
    db = SessionLocal()
    resp = admin_users(request=req_admin, filter="all", sort_key="role", sort_dir="asc", db=db, admin=admin_user)
    users_sorted = resp.context["users"]
    
    active_sorted = [u.user_id for u in users_sorted if u.is_active and u.user_id in temp_uids]
    
    # 활성 유저 순서 검증
    expected_active = ["t_admin", "t_pm", "t_lead", "t_staff1", "t_staff2"]
    assert active_sorted == expected_active, f"기본 정렬 활성 사용자 순서 오류: {active_sorted} != {expected_active}"
    
    # 비활성 유저는 무조건 리스트의 맨 마지막 그룹이어야 함
    last_uids = [u.user_id for u in users_sorted[-2:]]
    assert "t_inactive1" in last_uids and "t_inactive2" in last_uids, f"비활성 사용자가 맨 아래에 오지 않음: {last_uids}"
    
    # 3. 역순 정렬 검증 (sort_key="role", sort_dir="desc")
    resp_desc = admin_users(request=req_admin, filter="all", sort_key="role", sort_dir="desc", db=db, admin=admin_user)
    users_sorted_desc = resp_desc.context["users"]
    
    active_sorted_desc = [u.user_id for u in users_sorted_desc if u.is_active and u.user_id in temp_uids]
    expected_active_desc = ["t_staff2", "t_staff1", "t_lead", "t_pm", "t_admin"]
    assert active_sorted_desc == expected_active_desc, f"역순 정렬 활성 사용자 순서 오류: {active_sorted_desc} != {expected_active_desc}"
    
    # 역순일 때도 비활성 사원은 여전히 맨 뒤에 있어야 함
    last_uids_desc = [u.user_id for u in users_sorted_desc[-2:]]
    assert "t_inactive1" in last_uids_desc and "t_inactive2" in last_uids_desc, f"역순 정렬 시 비활성 사용자가 맨 아래에 오지 않음: {last_uids_desc}"
    
    # 4. 다른 정렬 키 검증 시 비활성 최하단 고정 검증 (예: sort_key="user_name", sort_dir="asc")
    resp_name = admin_users(request=req_admin, filter="all", sort_key="user_name", sort_dir="asc", db=db, admin=admin_user)
    users_sorted_name = resp_name.context["users"]
    
    active_sorted_name = [u.user_id for u in users_sorted_name if u.is_active and u.user_id in temp_uids]
    expected_active_name = ["t_admin", "t_pm", "t_lead", "t_staff1", "t_staff2"]
    assert active_sorted_name == expected_active_name, f"이름 정렬 활성 사용자 순서 오류: {active_sorted_name} != {expected_active_name}"
    
    last_uids_name = [u.user_id for u in users_sorted_name[-2:]]
    assert "t_inactive1" in last_uids_name and "t_inactive2" in last_uids_name, f"이름 정렬 시 비활성 사용자가 맨 아래에 오지 않음: {last_uids_name}"
    
    # 5. 테스트 유저 정리
    db.query(models.Users).filter(models.Users.user_id.in_(temp_uids)).delete(synchronize_session=False)
    db.commit()
    db.close()
    print("  -> PASS: User sorting & inactive isolation logic verified successfully.")

    # --- 시나리오 21: 비밀키 일관성 검증 및 Fail-Fast 구동 차단 검증 (v1.6.1) ---
    print("[CASE 21] Secret Key Consistency & Fail-Fast Verification")
    import hashlib
    from sqlalchemy import text
    
    db = SessionLocal()
    settings = db.query(models.SystemSettings).first()
    assert settings is not None
    # 0. 원래의 스냅샷 값을 백업해둠
    original_snapshot = settings.key_hash_snapshot
    db.close()
    
    try:
        # A. 평문 모드로 DB가 이미 기동되어 설정된 상황을 모사
        # 1. 스냅샷 값을 'PLAINTEXT_MODE'로 업데이트
        db = SessionLocal()
        db.execute(text("UPDATE system_settings SET key_hash_snapshot = 'PLAINTEXT_MODE'"))
        db.commit()
        db.close()
        
        # 환경변수 클리어 (평문 모드로 대조 실행 준비)
        os.environ.pop("SHIM_SECRET_KEY", None)
        auth.get_encryption_key.cache_clear()
        
        # 평문 기동 시도 -> 성공해야 함
        startup_event()
        
        # 2. 평문인 상태의 DB에 임의의 암호키 주입 후 기동 시도 -> Fail-Fast 차단 검증
        os.environ["SHIM_SECRET_KEY"] = "test_temp_secure_key_12345"
        auth.get_encryption_key.cache_clear()
        try:
            startup_event()
            assert False, "평문 DB에 키를 주입했는데 기동 차단이 되지 않았습니다."
        except SystemExit as se:
            assert se.code == 1, "정상적으로 sys.exit(1)로 차단됨"
            
        # B. 암호화 모드로 DB가 이미 기동되어 설정된 상황을 모사
        # 1. 스냅샷 값을 'key_aaa'의 해시값으로 업데이트
        os.environ["SHIM_SECRET_KEY"] = "key_aaa"
        auth.get_encryption_key.cache_clear()
        key_hash_aaa = hashlib.sha256(auth.get_encryption_key()).hexdigest()
        
        db = SessionLocal()
        db.execute(text("UPDATE system_settings SET key_hash_snapshot = :khash"), {"khash": key_hash_aaa})
        db.commit()
        db.close()
        
        # 암호화 기동 시도 -> 성공해야 함
        startup_event()
        
        # 2. 암호화된 DB에 키 주입 누락 후 기동 시도 -> 차단 검증
        os.environ.pop("SHIM_SECRET_KEY", None)
        auth.get_encryption_key.cache_clear()
        try:
            startup_event()
            assert False, "암호화 DB에 키가 누락되었는데 기동 차단이 되지 않았습니다."
        except SystemExit as se:
            assert se.code == 1, "정상적으로 sys.exit(1)로 차단됨"
            
        # 3. 암호화된 DB에 다른 키 주입 후 기동 시도 -> 차단 검증
        os.environ["SHIM_SECRET_KEY"] = "key_bbb"
        auth.get_encryption_key.cache_clear()
        try:
            startup_event()
            assert False, "암호화 DB에 다른 키가 주입되었는데 기동 차단이 되지 않았습니다."
        except SystemExit as se:
            assert se.code == 1, "정상적으로 sys.exit(1)로 차단됨"
            
    finally:
        # 원래 상태로 데이터베이스 복구
        db = SessionLocal()
        db.execute(text("UPDATE system_settings SET key_hash_snapshot = :orig"), {"orig": original_snapshot})
        db.commit()
        db.close()
        os.environ.pop("SHIM_SECRET_KEY", None)
        auth.get_encryption_key.cache_clear()
        
    print("  -> PASS: Key consistency and fail-fast blocking verified.")

    # --- 시나리오 22: PII 투명 개인정보 컬럼 암복호화 검증 ---
    print("[CASE 22] PII Column Transparent Encryption & Decryption")
    # 임시 키 설정 (암호화 모드)
    os.environ["SHIM_SECRET_KEY"] = "test_crypt_secret_key_for_pii_check_5678"
    auth.get_encryption_key.cache_clear()
    
    # DB 스냅샷 강제 세팅 (기동 차단 우회)
    db = SessionLocal()
    import hashlib
    key_hash = hashlib.sha256(auth.get_encryption_key()).hexdigest()
    
    settings = db.query(models.SystemSettings).first()
    original_snapshot = settings.key_hash_snapshot if settings else None
    
    db.execute(text("UPDATE system_settings SET key_hash_snapshot = :khash"), {"khash": key_hash})
    db.commit()
    db.close()
    
    # u_crypt 일반 사원 생성 및 연차 신청
    # (user_name, Leaves.reason, Leaves.rejection_reason은 EncryptedString 타입임)
    db = SessionLocal()
    # 청소
    db.query(models.Leaves).filter(models.Leaves.user_id == "u_crypt").delete()
    db.query(models.Users).filter(models.Users.user_id == "u_crypt").delete()
    db.commit()
    
    db.add(models.Users(
        user_id="u_crypt", user_name="홍길동", role="STAFF",
        password=auth.get_password_hash("0000"), is_active=True
    ))
    db.commit()
    
    # 연차 신청 추가
    db.add(models.Leaves(
        user_id="u_crypt", date=date(2026, 12, 25), snapshot_slot_label="09:00~18:00",
        snapshot_start_min=540, snapshot_end_min=1080, snapshot_deduction_hours=8.0,
        status="REJECTED", year=2026, reason="크리스마스 휴무", rejection_reason="업무 지원 필요"
    ))
    db.commit()
    db.close()
    
    # 검증 1) ORM 객체를 통해 읽었을 때 평문 복호화 확인
    db = SessionLocal()
    u_obj = db.query(models.Users).filter(models.Users.user_id == "u_crypt").first()
    assert u_obj.user_name == "홍길동", f"ORM 복호화 실패: {u_obj.user_name}"
    
    l_obj = db.query(models.Leaves).filter(models.Leaves.user_id == "u_crypt").first()
    assert l_obj.reason == "크리스마스 휴무", f"ORM 복호화 실패: {l_obj.reason}"
    assert l_obj.rejection_reason == "업무 지원 필요", f"ORM 복호화 실패: {l_obj.rejection_reason}"
    
    # 검증 2) RAW SQL 실행하여 DB 파일상 실제로 암호화된 토큰 형태(gAAAAAB로 시작)로 들어갔는지 검사
    raw_user_name = db.execute(text("SELECT user_name FROM users WHERE user_id = 'u_crypt'")).scalar()
    assert raw_user_name.startswith("gAAAAAB"), f"DB 내부 암호화 확인 실패: {raw_user_name}"
    
    raw_leave_reason = db.execute(text("SELECT reason FROM leaves WHERE user_id = 'u_crypt'")).scalar()
    assert raw_leave_reason.startswith("gAAAAAB"), f"DB 내부 암호화 확인 실패: {raw_leave_reason}"
    
    raw_reject_reason = db.execute(text("SELECT rejection_reason FROM leaves WHERE user_id = 'u_crypt'")).scalar()
    assert raw_reject_reason.startswith("gAAAAAB"), f"DB 내부 암호화 확인 실패: {raw_reject_reason}"
    
    # 청소
    db.query(models.Leaves).filter(models.Leaves.user_id == "u_crypt").delete()
    db.query(models.Users).filter(models.Users.user_id == "u_crypt").delete()
    # 스냅샷 복구
    db.execute(text("UPDATE system_settings SET key_hash_snapshot = :orig"), {"orig": original_snapshot})
    db.commit()
    db.close()
    os.environ.pop("SHIM_SECRET_KEY", None)
    auth.get_encryption_key.cache_clear()
    
    print("  -> PASS: PII Column Transparent Encryption & Decryption verified.")

    # --- 시나리오 23: 라우터 권한 검사 (Depends 재귀 검사) ---
    print("[CASE 23] Router Permission Gate Leak Audit")
    from fastapi.routing import APIRoute
    from fastapi.dependencies.utils import get_dependant
    from src.app.dependencies import get_current_admin
    
    def collect_all_dependencies(route: APIRoute) -> set:
        dependencies = set()
        if hasattr(route, "dependencies") and route.dependencies:
            for dep in route.dependencies:
                if hasattr(dep, "dependency") and dep.dependency:
                    dependencies.add(dep.dependency)
                elif hasattr(dep, "call") and dep.call:
                    dependencies.add(dep.call)
        if hasattr(route, "dependant") and route.dependant:
            def recurse_deps(dep_obj):
                if hasattr(dep_obj, "call") and dep_obj.call:
                    dependencies.add(dep_obj.call)
                if hasattr(dep_obj, "dependency") and dep_obj.dependency:
                    dependencies.add(dep_obj.dependency)
                if hasattr(dep_obj, "dependencies") and dep_obj.dependencies:
                    for child in dep_obj.dependencies:
                        recurse_deps(child)
            recurse_deps(route.dependant)
        return dependencies

    exempt_paths = {
        "/admin/login",
    }
    leaked_routes = []
    for route in app.routes:
        if not isinstance(route, APIRoute):
            continue
        path = route.path
        if path.startswith(("/admin", "/api/admin")) and path not in exempt_paths:
            all_deps = collect_all_dependencies(route)
            if get_current_admin not in all_deps:
                leaked_routes.append(f"[{route.methods}] {path} -> {route.endpoint.__name__}")
                
    assert not leaked_routes, f"관리자 권한 가드 누락 엔드포인트 발견: {leaked_routes}"
    print("  -> PASS: Router permission leak audit completed successfully.")

    # --- 시나리오 24: CORS Preflight 테스트 ---
    print("[CASE 24] Closed Network CORS Subnet & Wildcard Matcher")
    from src.app.middlewares.cors import ClosedNetworkCORSMiddleware
    from fastapi import FastAPI
    
    test_cors_app = FastAPI()
    test_cors_app.add_middleware(
        ClosedNetworkCORSMiddleware,
        origins_raw="http://*.local,http://192.168.10.0/24"
    )
    @test_cors_app.get("/api/ping")
    def ping():
        return {"ping": "pong"}
        
    cors_client = TestClient(test_cors_app)
    
    # 1. 와일드카드 도메인 통과
    r1 = cors_client.options("/api/ping", headers={
        "Origin": "http://shim-service.local",
        "Access-Control-Request-Method": "GET"
    })
    assert r1.status_code == 204
    assert r1.headers.get("access-control-allow-origin") == "http://shim-service.local"
    assert r1.headers.get("access-control-allow-credentials") == "true"
    
    # 2. CIDR 서브넷 대역 통과
    r2 = cors_client.options("/api/ping", headers={
        "Origin": "http://192.168.10.15",
        "Access-Control-Request-Method": "GET"
    })
    assert r2.status_code == 204
    assert r2.headers.get("access-control-allow-origin") == "http://192.168.10.15"

    # 3. 비허용 대역 차단
    r3 = cors_client.options("/api/ping", headers={
        "Origin": "http://192.168.99.15",
        "Access-Control-Request-Method": "GET"
    })
    assert "access-control-allow-origin" not in r3.headers
    print("  -> PASS: Subnet & wildcard CORS preflight verified.")



    # --- 시나리오 26: DBInitLock 동시성 락 ---
    print("[CASE 26] DBInitLock Atomic Directory Concurrency Guard")
    from tools.scripts.db_init import DBInitLock
    
    lock_dir = TEST_DATA_DIR / "test_migration.lock"
    # 깨끗하게 초기화
    if lock_dir.exists():
        if (lock_dir / "lock.time").exists():
            (lock_dir / "lock.time").unlink()
        lock_dir.rmdir()
        
    # 1. 락 획득 성공 검증
    with DBInitLock(lock_dir) as lock1:
        assert lock_dir.exists()
        assert (lock_dir / "lock.time").exists()
        
        # 2. 획득 중인 상태에서 중복 획득 시도 시 타임아웃 예외
        try:
            with DBInitLock(lock_dir, timeout=1):
                assert False, "이미 락을 쥔 상태이므로 예외가 발생해야 합니다."
        except TimeoutError:
            pass  # 정상 작동
            
    # 3. 락 해제 후 정상 삭제되었는지 검증
    assert not lock_dir.exists()
    print("  -> PASS: DBInitLock atomic directory locking verified.")

    # --- 시나리오 27: 알림 폴링 및 읽음 처리 API 검증 ---
    print("[CASE 27] Notification Polling & Read API Verification")
    db = SessionLocal()
    # 청소
    db.query(models.Notifications).filter(models.Notifications.user_id == "u_staff").delete()
    db.commit()
    db.close()

    # 1. 알림 생성
    db = SessionLocal()
    utils.create_notification(db, user_id="u_staff", sender_id="u_pm", message="테스트 알림 메시지 1")
    db.commit()
    db.close()

    # 2. 알림 조회 (폴링)
    r_poll = staff_client.get("/api/notifications")
    assert r_poll.status_code == 200
    notifications = r_poll.json()
    assert len(notifications) == 1
    assert notifications[0]["message"] == "테스트 알림 메시지 1"
    assert notifications[0]["is_read"] is False
    noti_id = notifications[0]["id"]

    # 3. 알림 읽음 처리
    r_read = staff_client.post(f"/api/notifications/{noti_id}/read")
    assert r_read.status_code == 200
    assert r_read.json() == {"status": "success"}

    # 4. 재조회 시 읽음 반영 확인
    r_poll2 = staff_client.get("/api/notifications")
    assert r_poll2.status_code == 200
    assert len(r_poll2.json()) == 0

    # 5. 모두 읽기 검증
    db = SessionLocal()
    utils.create_notification(db, user_id="u_staff", sender_id="u_pm", message="테스트 알림 메시지 2")
    utils.create_notification(db, user_id="u_staff", sender_id="u_pm", message="테스트 알림 메시지 3")
    db.commit()
    db.close()

    r_poll3 = staff_client.get("/api/notifications")
    assert len(r_poll3.json()) == 2

    r_read_all = staff_client.post("/api/notifications/read-all")
    assert r_read_all.status_code == 200
    assert r_read_all.json() == {"status": "success"}

    r_poll4 = staff_client.get("/api/notifications")
    assert len(r_poll4.json()) == 0
    print("  -> PASS: Notification polling and reading APIs verified.")

    # --- 시나리오 28: 정기 알림 삭제(Cleanup) 및 30일 경과 필터 검증 ---
    print("[CASE 28] Notification Cleanup Scheduler & 30-Day Age Limit")
    from src.app.services import ops
    
    db = SessionLocal()
    db.query(models.Notifications).filter(models.Notifications.user_id == "u_staff").delete()
    db.commit()
    db.close()

    # 오래된 알림(31일 전)과 최근 알림(5일 전) 생성
    db = SessionLocal()
    old_noti = models.Notifications(
        user_id="u_staff",
        sender_id="u_pm",
        message="31일 전 오래된 알림",
        created_at=datetime.utcnow() - timedelta(days=31)
    )
    new_noti = models.Notifications(
        user_id="u_staff",
        sender_id="u_pm",
        message="5일 전 최근 알림",
        created_at=datetime.utcnow() - timedelta(days=5)
    )
    db.add(old_noti)
    db.add(new_noti)
    db.commit()
    
    old_id = old_noti.id
    new_id = new_noti.id
    db.close()

    # 클린업 수행
    ops.cleanup_old_notifications()

    # 검증: 오래된 알림은 지워지고 최근 알림은 남아 있어야 함
    db = SessionLocal()
    remaining_old = db.query(models.Notifications).filter(models.Notifications.id == old_id).first()
    remaining_new = db.query(models.Notifications).filter(models.Notifications.id == new_id).first()
    
    assert remaining_old is None, "31일 경과된 오래된 알림이 삭제되지 않았습니다."
    assert remaining_new is not None, "5일 경과된 최근 알림이 삭제되었습니다."
    assert remaining_new.message == "5일 전 최근 알림"
    
    # 청소
    db.query(models.Notifications).filter(models.Notifications.user_id == "u_staff").delete()
    db.commit()
    db.close()
    print("  -> PASS: Notification 30-day age cleanup verified.")

    print("\n[COMPLETE] All key features verified successfully.")
    db = SessionLocal()
    db.close()

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n[FAILURE] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
(1)
