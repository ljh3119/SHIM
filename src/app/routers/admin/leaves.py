from datetime import datetime, date as date_cls, timedelta
from urllib.parse import urlencode
import calendar
import io

from fastapi import APIRouter, Depends, Request, Form, status
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, RedirectResponse
from sqlalchemy import extract, func
from sqlalchemy.orm import Session, contains_eager
from sqlalchemy.exc import SQLAlchemyError
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell

from src.app import models, database, auth, utils
from src.app.database import get_db
from src.app.dependencies import get_current_admin
from src.app.services import admin_service
from src.app.services.leave_service import resolve_user_yearly_allocated_hours
from src.app.services.leave_policy import (
    apply_leave_status_transition,
    LeaveStatusTransitionError,
)

page_router = APIRouter()
api_router = APIRouter()

def _templates(request: Request):
    return request.app.state.templates

TIMELINE_SORT_COLUMNS = frozenset({"created_at", "user_name", "date", "slot", "company", "team"})
CALENDAR_SORT_COLUMNS = frozenset({"user_name", "company", "team", "yearly_remain", "period_used"})
TIMELINE_LEAVE_STATUS_FILTERS = frozenset({"PENDING", "APPROVED", "REJECTED"})

def _calendar_next_sort_dir(column: str, current_sort: str | None, current_dir: str) -> str:
    if current_sort != column:
        if column in ("user_name", "company", "team"):
            return "asc"
        return "desc"
    return "desc" if current_dir == "asc" else "asc"

def _timeline_next_sort_dir(column: str, current_sort: str, current_dir: str) -> str:
    if current_sort != column:
        return "desc" if column in ("created_at", "date") else "asc"
    return "asc" if current_dir == "desc" else "desc"

def _timeline_leave_status_filter(raw: str | None) -> str:
    s = (raw or "").strip().upper()
    return s if s in TIMELINE_LEAVE_STATUS_FILTERS else ""


@page_router.get("/leave/timeline")
def admin_leaves_timeline(
    request: Request,
    admin: models.Users = Depends(get_current_admin),
):
    # Redirect to calendar page with timeline tab active, keeping query params
    params = dict(request.query_params)
    params["tab"] = "timeline"
    return RedirectResponse(url=f"/admin/leave/calendar?{urlencode(params)}")


@page_router.get("/leave/timeline/partial", response_class=HTMLResponse)
def admin_leaves_timeline_partial(
    request: Request,
    year: int = None,
    month: int = 0,
    user_id: str = "",
    company: str = "",
    team: str = "",
    leave_status: str = "",
    sort: str = "created_at",
    sort_dir: str = "desc",
    order: str | None = None,
    page: int = 1,
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
    current_year = year if year else utils.get_local_now().year
    current_month = month
    selected_user_id = user_id.strip()
    selected_company = company.strip()
    selected_team = team.strip()
    selected_leave_status = _timeline_leave_status_filter(leave_status)

    sort_key = sort if sort in TIMELINE_SORT_COLUMNS else "created_at"
    if "sort" not in request.query_params and order in ("asc", "desc"):
        sort_dir = order
    sort_dir_effective = sort_dir if sort_dir in ("asc", "desc") else "desc"

    leave_year_rows = db.query(models.Leaves.year).distinct().all()
    leave_years = [row[0] for row in leave_year_rows]
    year_options = utils.build_year_options(utils.get_local_now().year, leave_years)

    query = admin_service.get_leaves_timeline_query(
        db=db,
        year=current_year,
        month=current_month,
        user_id=selected_user_id,
        company=selected_company,
        team=selected_team,
        leave_status=selected_leave_status,
    )
    query = query.filter(extract("year", models.Leaves.date) == current_year)

    per_page = 50
    total_count = query.count()
    is_capped = total_count >= 5000
    total_pages = (total_count + per_page - 1) // per_page
    if page > total_pages and total_pages > 0:
        page = total_pages
    if page < 1:
        page = 1

    sort_col = {
        "created_at": models.Leaves.created_at,
        "user_name": models.Users.user_name,
        "date": models.Leaves.date,
        "slot": models.Leaves.snapshot_slot_label,
        "company": models.Users.company,
        "team": models.Users.team,
    }[sort_key]

    is_encrypted = auth.get_encryption_key() is not None
    if is_encrypted and sort_key == "user_name":
        # If encryption is active, DB-level sorting of user_name is meaningless (as it is encrypted).
        # We query all rows, sort in memory by the decrypted user_name (using lower case), and then apply pagination.
        all_leaves = query.limit(5000).all()
        reverse_sort = (sort_dir_effective == "desc")
        all_leaves.sort(
            key=lambda x: (
                (x.user.user_name or "").lower(),
                x.id if not reverse_sort else -x.id
            ),
            reverse=reverse_sort
        )
        leaves = all_leaves[(page - 1) * per_page : page * per_page]
    else:
        if sort_dir_effective == "asc":
            query = query.order_by(sort_col.asc().nulls_last(), models.Leaves.id.asc())
        else:
            query = query.order_by(sort_col.desc().nulls_last(), models.Leaves.id.desc())
        leaves = query.offset((page - 1) * per_page).limit(per_page).all()

    all_users = (
        db.query(models.Users)
        .filter(models.Users.role != "ADMIN")
        .all()
    )
    users = sorted(all_users, key=lambda u: u.user_name.lower())
    company_options = sorted(list({u.company for u in all_users if u.company}))
    team_options = sorted(list({u.team for u in all_users if u.team}))

    path = request.url.path
    base_q = {"year": str(current_year), "month": str(current_month)}
    if selected_user_id:
        base_q["user_id"] = selected_user_id
    if selected_company:
        base_q["company"] = selected_company
    if selected_team:
        base_q["team"] = selected_team
    if selected_leave_status:
        base_q["leave_status"] = selected_leave_status

    sort_urls = {}
    for col in TIMELINE_SORT_COLUMNS:
        nxt = _timeline_next_sort_dir(col, sort_key, sort_dir_effective)
        q = {**base_q, "sort": col, "sort_dir": nxt}
        sort_urls[col] = f"{path}?{urlencode(q)}"

    chip_q_base = {"year": str(current_year), "month": str(current_month)}
    if selected_user_id:
        chip_q_base["user_id"] = selected_user_id
    if selected_company:
        chip_q_base["company"] = selected_company
    if selected_team:
        chip_q_base["team"] = selected_team

    def _timeline_chip_url(extra: dict) -> str:
        q = {
            **chip_q_base,
            "sort": sort_key,
            "sort_dir": sort_dir_effective,
            **extra,
        }
        q = {k: v for k, v in q.items() if v not in ("", None)}
        return f"{path}?{urlencode(q)}"

    timeline_chip_url_all = _timeline_chip_url({})
    timeline_chip_url_pending = _timeline_chip_url({"leave_status": "PENDING"})
    timeline_chip_url_approved = _timeline_chip_url({"leave_status": "APPROVED"})
    timeline_chip_url_rejected = _timeline_chip_url({"leave_status": "REJECTED"})

    return _templates(request).TemplateResponse(
        request=request,
        name="partials/admin_leaves_timeline_partial.html",
        context={
            "admin": admin,
            "leaves": leaves,
            "users": users,
            "is_capped": is_capped,
            "company_options": company_options,
            "team_options": team_options,
            "selected_year": current_year,
            "selected_month": current_month,
            "selected_user_id": selected_user_id,
            "selected_company": selected_company,
            "selected_team": selected_team,
            "sort_key": sort_key,
            "sort_dir": sort_dir_effective,
            "sort_urls": sort_urls,
            "page": page,
            "total_pages": total_pages,
            "total_count": total_count,
            "export_query": urlencode(
                {
                    k: v
                    for k, v in {
                        "year": str(current_year),
                        "month": str(current_month),
                        "user_id": selected_user_id,
                        "company": selected_company,
                        "team": selected_team,
                        "leave_status": selected_leave_status,
                        "sort": sort_key,
                        "sort_dir": sort_dir_effective,
                    }.items()
                    if v not in ("", None)
                }
            ),
            "selected_leave_status": selected_leave_status,
            "timeline_chip_url_all": timeline_chip_url_all,
            "timeline_chip_url_pending": timeline_chip_url_pending,
            "timeline_chip_url_approved": timeline_chip_url_approved,
            "timeline_chip_url_rejected": timeline_chip_url_rejected,
            "current_year": utils.get_local_now().year,
            "year_options": year_options,
        },
    )


@page_router.get("/leave/timeline/export")
def export_admin_leaves_timeline(
    request: Request,
    year: int = None,
    month: int = 0,
    user_id: str = "",
    company: str = "",
    team: str = "",
    leave_status: str = "",
    sort: str = "created_at",
    sort_dir: str = "desc",
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
    current_year = year if year else utils.get_local_now().year
    selected_user_id = user_id.strip()
    selected_company = company.strip()
    selected_team = team.strip()
    selected_leave_status = _timeline_leave_status_filter(leave_status)
    sort_key = sort if sort in TIMELINE_SORT_COLUMNS else "created_at"
    sort_dir_effective = sort_dir if sort_dir in ("asc", "desc") else "desc"

    query = (
        db.query(models.Leaves)
        .join(models.Users, models.Leaves.user_id == models.Users.user_id)
        .options(contains_eager(models.Leaves.user))
        .filter(models.Leaves.year == current_year)
    )
    if month > 0:
        import calendar as cal_module
        num_days = cal_module.monthrange(current_year, month)[1]
        query = query.filter(
            models.Leaves.date >= date_cls(current_year, month, 1),
            models.Leaves.date <= date_cls(current_year, month, num_days)
        )
    if selected_user_id:
        query = query.filter(models.Leaves.user_id == selected_user_id)
    if selected_company:
        query = query.filter(models.Users.company == selected_company)
    if selected_team:
        query = query.filter(models.Users.team == selected_team)
    if selected_leave_status:
        query = query.filter(models.Leaves.status == selected_leave_status)

    sort_col = {
        "created_at": models.Leaves.created_at,
        "user_name": models.Users.user_name,
        "date": models.Leaves.date,
        "slot": models.Leaves.snapshot_slot_label,
        "company": models.Users.company,
        "team": models.Users.team,
    }[sort_key]
    if sort_dir_effective == "asc":
        query = query.order_by(sort_col.asc().nulls_last(), models.Leaves.id.asc())
    else:
        query = query.order_by(sort_col.desc().nulls_last(), models.Leaves.id.desc())
    leaves = query.all()

    wb = Workbook(write_only=True)
    
    user_query = db.query(models.Users)
    if selected_user_id:
        user_query = user_query.filter(models.Users.user_id == selected_user_id)
    if selected_company:
        user_query = user_query.filter(models.Users.company == selected_company)
    if selected_team:
        user_query = user_query.filter(models.Users.team == selected_team)
    
    users = user_query.order_by(models.Users.company.asc(), models.Users.team.asc(), models.Users.user_name.asc()).all()
    user_ids = [u.user_id for u in users]

    allocations = (
        db.query(models.UserYearlyLeaveAllocations)
        .filter(
            models.UserYearlyLeaveAllocations.user_id.in_(user_ids),
            models.UserYearlyLeaveAllocations.year == current_year
        )
        .all()
    )
    alloc_map = {a.user_id: a.allocated_hours for a in allocations}

    leaves_summary = (
        db.query(
            models.Leaves.user_id,
            models.Leaves.status,
            func.sum(models.Leaves.snapshot_deduction_hours).label("total_hours")
        )
        .filter(
            models.Leaves.user_id.in_(user_ids),
            models.Leaves.year == current_year,
            models.Leaves.is_deductive == True
        )
        .group_by(models.Leaves.user_id, models.Leaves.status)
        .all()
    )

    leave_hours_map = {uid: {"APPROVED": 0.0, "PENDING": 0.0} for uid in user_ids}
    for uid, status_val, total_hours in leaves_summary:
        if uid in leave_hours_map and status_val in ("APPROVED", "PENDING"):
            leave_hours_map[uid][status_val] = float(total_hours or 0.0)

    ws_summary = wb.create_sheet(title="연차 현황 요약")
    ws_summary.append([
        "사원 ID", "사원명", "회사", "팀", "직급", "역할", "계정 상태",
        f"{current_year}년 총 배정 시간 (h)", "사용 완료 시간 (h)", "결재 대기 시간 (h)", "잔여 시간 (h)"
    ])

    for u in users:
        allocated = alloc_map.get(u.user_id, u.total_leave_hours or 120)
        approved = leave_hours_map.get(u.user_id, {}).get("APPROVED", 0.0)
        pending = leave_hours_map.get(u.user_id, {}).get("PENDING", 0.0)
        remaining = float(allocated) - approved - pending
        
        ws_summary.append([
            u.user_id,
            u.user_name,
            u.company or "",
            u.team or "",
            u.position or "",
            u.role or "",
            "활성" if u.is_active else "비활성",
            float(allocated),
            float(approved),
            float(pending),
            float(remaining)
        ])

    ws_timeline = wb.create_sheet(title="상세 신청 내역")
    ws_timeline.append([
        "신청 시각", "사용자명", "사용자 ID", "회사", "팀", "연차일", 
        "사용 시간대", "차감 여부", "차감시간 (h)", "상태", "신청 사유", "반려 사유"
    ])

    for leave in leaves:
        row = []
        if leave.created_at:
            cell_created = WriteOnlyCell(ws_timeline, value=utils.to_kst_naive(leave.created_at))
            cell_created.number_format = 'yyyy-mm-dd hh:mm:ss'
            row.append(cell_created)
        else:
            row.append("")
            
        row.append(leave.user.user_name if leave.user else "")
        row.append(leave.user.user_id if leave.user else "")
        row.append((leave.user.company or "") if leave.user else "")
        row.append((leave.user.team or "") if leave.user else "")
        
        if leave.date:
            cell_date = WriteOnlyCell(ws_timeline, value=leave.date)
            cell_date.number_format = 'yyyy-mm-dd'
            row.append(cell_date)
        else:
            row.append("")
            
        row.append(leave.snapshot_slot_label or "")
        row.append("차감" if leave.is_deductive else "비차감")
        row.append(float(leave.snapshot_deduction_hours or 0.0))
        row.append(leave.status or "")
        row.append(leave.reason or "")
        row.append(leave.rejection_reason or "")
        
        ws_timeline.append(row)

    output = io.BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    stamp = utils.get_local_now().strftime("%Y%m%d_%H%M%S")
    filename = f"leave_summary_{current_year}_{stamp}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@page_router.get("/leave/calendar", response_class=HTMLResponse)
def admin_leaves_calendar(
    request: Request,
    year: int = None,
    month: int = None,
    view: str = "month",
    user_id: str = "",
    company: str = "",
    team: str = "",
    active_state: str = "all",
    sort: str = "",
    sort_dir: str = "asc",
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
    current_year = year if year else utils.get_local_now().year
    current_month = month if month else utils.get_local_now().month
    if view not in ("month", "month_grid", "year"):
        view = "month"
    is_year_view = view == "year"
    selected_user_id = user_id.strip()
    selected_company = company.strip()
    selected_team = team.strip()
    selected_active_state = active_state if active_state in ("all", "active", "inactive") else "all"
    sort_dir_eff = sort_dir if sort_dir in ("asc", "desc") else "asc"
    sort_param = (sort or "").strip()
    sort_key_effective = sort_param if sort_param in CALENDAR_SORT_COLUMNS else None

    leave_year_rows = db.query(models.Leaves.year).distinct().all()
    leave_years = [row[0] for row in leave_year_rows]
    year_options = utils.build_year_options(utils.get_local_now().year, leave_years)

    all_users = db.query(models.Users).filter(models.Users.role != "ADMIN").all()
    
    company_options = sorted(list({u.company for u in all_users if u.company}))
    team_options = sorted(list({u.team for u in all_users if u.team}))
    
    users_all = [
        u for u in all_users
        if selected_active_state == "all"
        or (selected_active_state == "active" and u.is_active)
        or (selected_active_state == "inactive" and not u.is_active)
    ]
    users_all.sort(key=lambda u: u.user_name.lower())
    
    users = [
        u for u in all_users
        if (selected_active_state == "all" or (selected_active_state == "active" and u.is_active) or (selected_active_state == "inactive" and not u.is_active))
        and (not selected_company or u.company == selected_company)
        and (not selected_team or u.team == selected_team)
        and (not selected_user_id or u.user_id == selected_user_id)
    ]

    weekday_labels = ["월", "화", "수", "목", "금", "토", "일"]

    user_leaves_map = {u.user_id: {} for u in users}
    days = []
    day_weekday_map = {}
    weekend_days = []
    holiday_day_map = {}
    user_month_hours = {u.user_id: {m: 0.0 for m in range(1, 13)} for u in users}

    user_ids = [u.user_id for u in users]
    user_yearly_leaves_map = {uid: [] for uid in user_ids}

    if user_ids:
        year_leaves = (
            db.query(models.Leaves)
            .filter(
                models.Leaves.user_id.in_(user_ids),
                models.Leaves.year == current_year,
                models.Leaves.status.notin_(["CANCELED", "REJECTED"]),
                models.Leaves.is_deductive == True
            )
            .all()
        )
        for l in year_leaves:
            user_yearly_leaves_map[l.user_id].append(l)

    offset = 0
    if is_year_view:
        for uid, leaves in user_yearly_leaves_map.items():
            for l in leaves:
                user_month_hours[uid][l.date.month] += float(l.snapshot_deduction_hours or 0)
    else:
        num_days = calendar.monthrange(current_year, current_month)[1]
        first_weekday_monday0 = calendar.weekday(current_year, current_month, 1)
        offset = (first_weekday_monday0 + 1) % 7
        days = list(range(1, num_days + 1))
        day_weekday_map = {
            d: weekday_labels[calendar.weekday(current_year, current_month, d)] for d in days
        }
        weekend_days = [d for d in days if calendar.weekday(current_year, current_month, d) >= 5]

        month_start = date_cls(current_year, current_month, 1)
        month_end = date_cls(current_year, current_month, num_days)
        month_holidays = (
            db.query(models.Holidays)
            .filter(models.Holidays.date >= month_start, models.Holidays.date <= month_end)
            .order_by(models.Holidays.date.asc())
            .all()
        )
        holiday_day_map = {h.date.day: h for h in month_holidays}

        month_leaves = []
        if user_ids:
            month_leaves = (
                db.query(models.Leaves)
                .filter(
                    models.Leaves.user_id.in_(user_ids),
                    models.Leaves.date >= month_start,
                    models.Leaves.date <= month_end,
                )
                .all()
            )
        for l in month_leaves:
            if l.user_id not in user_leaves_map:
                continue
            day = l.date.day
            if day not in user_leaves_map[l.user_id]:
                user_leaves_map[l.user_id][day] = []
            user_leaves_map[l.user_id][day].append(l)

    allocation_map = admin_service.get_yearly_allocation_map(db, [u.user_id for u in users], current_year)
    user_stats = []
    for u in users:
        if is_year_view:
            yearly_used = sum(user_month_hours[u.user_id].values())
            month_used_labels = {
                m: (
                    utils.hours_to_days_hours_label(user_month_hours[u.user_id][m])
                    if user_month_hours[u.user_id][m]
                    else "-"
                )
                for m in range(1, 13)
            }
            month_used_short_labels = {
                m: (
                    "-"
                    if not user_month_hours[u.user_id][m]
                    else utils.hours_to_days_hours_compact(user_month_hours[u.user_id][m])
                )
                for m in range(1, 13)
            }
        else:
            month_used_labels = {}
            month_used_short_labels = {}
            yearly_leaves = user_yearly_leaves_map.get(u.user_id, [])
            yearly_used = sum(float(l.snapshot_deduction_hours or 0) for l in yearly_leaves)

        allocated_hours = float(allocation_map.get(u.user_id, u.total_leave_hours or 0))
        yearly_remain = allocated_hours - float(yearly_used)

        if is_year_view:
            period_used = yearly_used
        else:
            period_used = 0.0
            if u.user_id in user_leaves_map:
                for dlist in user_leaves_map[u.user_id].values():
                    period_used += sum(float(lv.snapshot_deduction_hours or 0) for lv in dlist if lv.status not in ("CANCELED", "REJECTED") and lv.is_deductive == True)

        user_stats.append(
            {
                "user_id": u.user_id,
                "user_name": u.user_name,
                "company": u.company,
                "team": u.team,
                "role": u.role,
                "total_hours": allocated_hours,
                "period_used": period_used,
                "yearly_used": yearly_used,
                "yearly_remain": yearly_remain,
                "yearly_remain_label": utils.hours_to_days_hours_label(yearly_remain),
                "period_used_label": utils.hours_to_days_hours_label(period_used),
                "yearly_remain_short": utils.hours_to_days_hours_compact(yearly_remain),
                "period_used_short": utils.hours_to_days_hours_compact(period_used),
                "month_used_labels": month_used_labels,
                "month_used_short_labels": month_used_short_labels,
                "is_active": u.is_active,
            }
        )

    if sort_key_effective is None:
        user_stats.sort(
            key=lambda x: (
                -int(x["is_active"]),
                0 if x["role"] == "PM" else 1,
                (x["team"] or "").lower(),
                0 if x["role"] == "TEAM_LEAD" else (1 if x["role"] == "STAFF" else 2),
                x["user_name"].lower(),
                x["user_id"]
            )
        )
    else:
        rev = sort_dir_eff == "desc"
        if sort_key_effective == "user_name":
            user_stats.sort(
                key=lambda x: (x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )
        elif sort_key_effective == "company":
            user_stats.sort(
                key=lambda x: ((x["company"] or "").lower(), x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )
        elif sort_key_effective == "team":
            user_stats.sort(
                key=lambda x: ((x["team"] or "").lower(), x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )
        elif sort_key_effective == "yearly_remain":
            user_stats.sort(
                key=lambda x: (x["yearly_remain"], x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )
        elif sort_key_effective == "period_used":
            user_stats.sort(
                key=lambda x: (x["period_used"], x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )

    path = request.url.path
    view_q = view
    base_q = {
        "year": str(current_year),
        "month": str(current_month),
        "view": view_q,
    }
    if selected_user_id:
        base_q["user_id"] = selected_user_id
    if selected_company:
        base_q["company"] = selected_company
    if selected_team:
        base_q["team"] = selected_team
    if selected_active_state != "all":
        base_q["active_state"] = selected_active_state

    sort_urls = {}
    for col in CALENDAR_SORT_COLUMNS:
        nxt = _calendar_next_sort_dir(col, sort_key_effective, sort_dir_eff)
        q = {**base_q, "sort": col, "sort_dir": nxt}
        sort_urls[col] = f"{path}?{urlencode(q)}"

    return _templates(request).TemplateResponse(
        request=request,
        name="admin_leaves_calendar.html",
        context={
            "admin": admin,
            "users": users_all,
            "user_leaves_map": dict(user_leaves_map),
            "user_stats": user_stats,
            "days": days,
            "day_weekday_map": day_weekday_map,
            "weekend_days": weekend_days,
            "holiday_day_map": holiday_day_map,
            "is_year_view": is_year_view,
            "view": view,
            "offset": offset,
            "selected_year": current_year,
            "selected_month": current_month,
            "selected_user_id": selected_user_id,
            "selected_company": selected_company,
            "selected_team": selected_team,
            "selected_active_state": selected_active_state,
            "company_options": company_options,
            "team_options": team_options,
            "sort_key": sort_key_effective,
            "sort_dir": sort_dir_eff,
            "sort_urls": sort_urls,
            "current_year": utils.get_local_now().year,
            "year_options": year_options,
        },
    )


@api_router.post("/leave/delete")
def delete_leave(
    request: Request,
    leave_id: int = Form(...),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
    leave = db.query(models.Leaves).filter(models.Leaves.id == leave_id).first()
    if not leave:
        return JSONResponse(status_code=404, content={"message": "Leave not found"})
        
    actor_name = leave.user.user_name if leave.user else "알수없음"
    l_date = str(leave.date)
    l_status = str(leave.status)
    l_hours = str(leave.snapshot_deduction_hours)
    
    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="DELETE_LEAVE",
        target_info=f"Leave:{leave_id} ({actor_name}, {l_date})",
        old_data=f"Status:{l_status}, Date:{l_date}, Hours:{l_hours}",
        new_data="DELETED"
    )
    db.add(audit)
    db.delete(leave)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    return JSONResponse(status_code=200, content={"message": "연차가 성공적으로 삭제되었습니다."})


@api_router.post("/leave/update-status")
def update_leave_status(
    request: Request,
    leave_id: int = Form(...),
    status_value: str = Form(...),
    rejection_reason: str = Form(""),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
    leave = db.query(models.Leaves).filter(models.Leaves.id == leave_id).first()
    if not leave:
        return JSONResponse(status_code=404, content={"message": "Leave not found"})

    try:
        transition = apply_leave_status_transition(
            leave=leave,
            status_value=status_value,
            rejection_reason=rejection_reason,
        )
    except LeaveStatusTransitionError as exc:
        return JSONResponse(status_code=400, content={"message": str(exc)})

    leave = db.query(models.Leaves).filter(models.Leaves.id == leave_id).first()
    db.add(
        models.AuditLogs(
            actor_id=admin.user_id,
            action="UPDATE_LEAVE_STATUS",
            target_info=f"Leave:{leave_id} ({leave.user.user_name if leave else 'Unknown'}, {leave.date if leave else 'Unknown'})",
            old_data=transition.audit_old_data,
            new_data=transition.audit_new_data,
        )
    )
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    return JSONResponse(status_code=200, content={"message": "상태가 변경되었습니다."})


@api_router.post("/leave/update-type")
def update_leave_type(
    request: Request,
    leave_id: int = Form(...),
    is_deductive: bool = Form(...),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
    leave = db.query(models.Leaves).filter(models.Leaves.id == leave_id).first()
    if not leave:
        return JSONResponse(status_code=404, content={"message": "신청 건을 찾을 수 없습니다."})

    old_type = "연차(차감)" if leave.is_deductive else "공가/출장(비차감)"
    new_type = "연차(차감)" if is_deductive else "공가/출장(비차감)"

    if leave.is_deductive == is_deductive:
        return JSONResponse(status_code=400, content={"message": "이미 해당 유형입니다."})

    if is_deductive:
        total_allocated = resolve_user_yearly_allocated_hours(db, leave.user, leave.year)
        used_hours = db.query(func.sum(models.Leaves.snapshot_deduction_hours)).filter(
            models.Leaves.user_id == leave.user_id,
            models.Leaves.year == leave.year,
            models.Leaves.status.notin_(["CANCELED", "REJECTED"]),
            models.Leaves.is_deductive == True,
            models.Leaves.id != leave.id
        ).scalar() or 0.0
        
        if used_hours + leave.snapshot_deduction_hours > total_allocated:
            return JSONResponse(status_code=400, content={"message": "사용자의 잔여 연차가 부족하여 연차로 변경할 수 없습니다."})

    leave.is_deductive = is_deductive
    
    db.add(
        models.AuditLogs(
            actor_id=admin.user_id,
            action="UPDATE_LEAVE_TYPE",
            target_info=f"Leave:{leave_id} ({leave.user.user_name}, {leave.date})",
            old_data=f"is_deductive={not is_deductive} ({old_type})",
            new_data=f"is_deductive={is_deductive} ({new_type})",
        )
    )
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    return JSONResponse(status_code=200, content={"message": f"유형이 {new_type}로 변경되었습니다."})


@page_router.get("/leave/yearly/partial", response_class=HTMLResponse)
def admin_leaves_yearly_partial(
    request: Request,
    year: int = None,
    user_id: str = "",
    company: str = "",
    team: str = "",
    active_state: str = "all",
    sort: str = "",
    sort_dir: str = "asc",
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
    current_year = year if year else utils.get_local_now().year
    selected_user_id = user_id.strip()
    selected_company = company.strip()
    selected_team = team.strip()
    selected_active_state = active_state if active_state in ("all", "active", "inactive") else "all"
    sort_dir_eff = sort_dir if sort_dir in ("asc", "desc") else "asc"
    sort_param = (sort or "").strip()
    sort_key_effective = sort_param if sort_param in CALENDAR_SORT_COLUMNS else None

    leave_year_rows = db.query(models.Leaves.year).distinct().all()
    leave_years = [row[0] for row in leave_year_rows]
    year_options = utils.build_year_options(utils.get_local_now().year, leave_years)

    all_users = db.query(models.Users).filter(models.Users.role != "ADMIN").all()
    
    company_options = sorted(list({u.company for u in all_users if u.company}))
    team_options = sorted(list({u.team for u in all_users if u.team}))
    
    users = [
        u for u in all_users
        if (selected_active_state == "all" or (selected_active_state == "active" and u.is_active) or (selected_active_state == "inactive" and not u.is_active))
        and (not selected_company or u.company == selected_company)
        and (not selected_team or u.team == selected_team)
        and (not selected_user_id or u.user_id == selected_user_id)
    ]

    user_month_hours = {u.user_id: {m: 0.0 for m in range(1, 13)} for u in users}
    user_ids = [u.user_id for u in users]
    user_yearly_leaves_map = {uid: [] for uid in user_ids}

    if user_ids:
        year_leaves = (
            db.query(models.Leaves)
            .filter(
                models.Leaves.user_id.in_(user_ids),
                models.Leaves.year == current_year,
                models.Leaves.status.notin_(["CANCELED", "REJECTED"]),
                models.Leaves.is_deductive == True
            )
            .all()
        )
        for l in year_leaves:
            user_yearly_leaves_map[l.user_id].append(l)

    for uid, leaves in user_yearly_leaves_map.items():
        for l in leaves:
            user_month_hours[uid][l.date.month] += float(l.snapshot_deduction_hours or 0)

    allocation_map = admin_service.get_yearly_allocation_map(db, [u.user_id for u in users], current_year)
    user_stats = []
    for u in users:
        yearly_used = sum(user_month_hours[u.user_id].values())
        month_used_labels = {
            m: (
                utils.hours_to_days_hours_label(user_month_hours[u.user_id][m])
                if user_month_hours[u.user_id][m]
                else "-"
            )
            for m in range(1, 13)
        }
        month_used_short_labels = {
            m: (
                "-"
                if not user_month_hours[u.user_id][m]
                else utils.hours_to_days_hours_compact(user_month_hours[u.user_id][m])
            )
            for m in range(1, 13)
        }

        allocated_hours = float(allocation_map.get(u.user_id, u.total_leave_hours or 0))
        yearly_remain = allocated_hours - float(yearly_used)

        user_stats.append(
            {
                "user_id": u.user_id,
                "user_name": u.user_name,
                "company": u.company,
                "team": u.team,
                "role": u.role,
                "total_hours": allocated_hours,
                "period_used": yearly_used,
                "yearly_used": yearly_used,
                "yearly_remain": yearly_remain,
                "yearly_remain_label": utils.hours_to_days_hours_label(yearly_remain),
                "period_used_label": utils.hours_to_days_hours_label(yearly_used),
                "yearly_remain_short": utils.hours_to_days_hours_compact(yearly_remain),
                "period_used_short": utils.hours_to_days_hours_compact(yearly_used),
                "month_used_labels": month_used_labels,
                "month_used_short_labels": month_used_short_labels,
                "is_active": u.is_active,
            }
        )

    if sort_key_effective is None:
        user_stats.sort(
            key=lambda x: (
                -int(x["is_active"]),
                0 if x["role"] == "PM" else 1,
                (x["team"] or "").lower(),
                0 if x["role"] == "TEAM_LEAD" else (1 if x["role"] == "STAFF" else 2),
                x["user_name"].lower(),
                x["user_id"]
            )
        )
    else:
        rev = sort_dir_eff == "desc"
        if sort_key_effective == "user_name":
            user_stats.sort(
                key=lambda x: (x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )
        elif sort_key_effective == "company":
            user_stats.sort(
                key=lambda x: ((x["company"] or "").lower(), x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )
        elif sort_key_effective == "team":
            user_stats.sort(
                key=lambda x: ((x["team"] or "").lower(), x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )
        elif sort_key_effective == "yearly_remain":
            user_stats.sort(
                key=lambda x: (x["yearly_remain"], x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )
        elif sort_key_effective == "period_used":
            user_stats.sort(
                key=lambda x: (x["period_used"], x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )

    path = request.url.path
    base_q = {
        "year": str(current_year),
    }
    if selected_user_id:
        base_q["user_id"] = selected_user_id
    if selected_company:
        base_q["company"] = selected_company
    if selected_team:
        base_q["team"] = selected_team
    if selected_active_state != "all":
        base_q["active_state"] = selected_active_state

    sort_urls = {}
    for col in CALENDAR_SORT_COLUMNS:
        nxt = _calendar_next_sort_dir(col, sort_key_effective, sort_dir_eff)
        q = {**base_q, "sort": col, "sort_dir": nxt}
        sort_urls[col] = f"{path}?{urlencode(q)}"

    return _templates(request).TemplateResponse(
        request=request,
        name="partials/admin_leaves_yearly_partial.html",
        context={
            "admin": admin,
            "user_stats": user_stats,
            "users": sorted(all_users, key=lambda u: u.user_name.lower()),
            "selected_year": current_year,
            "selected_user_id": selected_user_id,
            "selected_company": selected_company,
            "selected_team": selected_team,
            "selected_active_state": selected_active_state,
            "company_options": company_options,
            "team_options": team_options,
            "sort_key": sort_key_effective,
            "sort_dir": sort_dir_eff,
            "sort_urls": sort_urls,
            "current_year": utils.get_local_now().year,
            "year_options": year_options,
        },
    )
