from __future__ import annotations

import calendar
import io
from datetime import datetime, date as date_cls, timedelta
from urllib.parse import urlencode

from fastapi import APIRouter, Depends, Request, Form, status, HTTPException, BackgroundTasks
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from sqlalchemy import extract, func
from sqlalchemy.orm import Session, contains_eager
from sqlalchemy.exc import SQLAlchemyError
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell

from .. import models, database, auth, utils
from ..database import get_db, DB_PATH
from ..services.leave_policy import (
    ALLOWED_TIME_GRANULARITIES,
    LeaveStatusTransitionError,
    apply_leave_status_transition,
    get_system_settings,
)
from ..services.ops import create_sqlite_backup, run_backup_and_rotate
from ..services import admin_service
from .api_user import resolve_user_yearly_allocated_hours

from ..dependencies import get_current_admin

page_router = APIRouter(prefix="/admin", tags=["admin_pages"])
api_router = APIRouter(prefix="/api/admin", tags=["admin_api"])


def _templates(request: Request):
    return request.app.state.templates


TIMELINE_SORT_COLUMNS = frozenset({"created_at", "user_name", "date", "slot", "company", "team"})

CALENDAR_SORT_COLUMNS = frozenset({"user_name", "company", "team", "yearly_remain", "period_used"})


def _calendar_next_sort_dir(column: str, current_sort: str | None, current_dir: str) -> str:
    if current_sort != column:
        if column in ("user_name", "company", "team"):
            return "asc"
        return "desc"
    return "desc" if current_dir == "asc" else "asc"


def _timeline_next_sort_dir(column: str, current_sort: str, current_dir: str) -> str:
    if current_sort != column:
        return "desc" if column in ("created_at", "date") else "asc"
    return "asc" if current_dir == "desc" else "desc"


TIMELINE_LEAVE_STATUS_FILTERS = frozenset({"PENDING", "APPROVED", "REJECTED"})


def _timeline_leave_status_filter(raw: str | None) -> str:
    s = (raw or "").strip().upper()
    return s if s in TIMELINE_LEAVE_STATUS_FILTERS else ""


def _ensure_system_setting(db: Session) -> models.SystemSettings:
    setting = db.query(models.SystemSettings).first()
    if not setting:
        setting = models.SystemSettings(
            is_approval_required=False,
            time_granularity_minutes=60,
            work_start_minute=9 * 60,
            work_end_minute=18 * 60,
            product_display_name="SHIM",
            product_nav_short="",
            brand_initial="S",
        )
        db.add(setting)
        db.commit()
        db.refresh(setting)
    if not getattr(setting, "time_granularity_minutes", None):
        setting.time_granularity_minutes = 60
    if not getattr(setting, "work_start_minute", None):
        setting.work_start_minute = 9 * 60
    if not getattr(setting, "work_end_minute", None):
        setting.work_end_minute = 18 * 60
    if setting.work_end_minute <= setting.work_start_minute:
        setting.work_start_minute = 9 * 60
        setting.work_end_minute = 18 * 60
    if setting.work_start_minute % 30 != 0:
        setting.work_start_minute = (setting.work_start_minute // 30) * 30
    if setting.work_end_minute % 30 != 0:
        setting.work_end_minute = ((setting.work_end_minute + 29) // 30) * 30
    if setting.work_end_minute > 24 * 60:
        setting.work_end_minute = 24 * 60
    if setting.work_end_minute <= setting.work_start_minute:
        setting.work_start_minute = 9 * 60
        setting.work_end_minute = 18 * 60
    db.commit()
    db.refresh(setting)
    return setting

@page_router.get("/dashboard", response_class=HTMLResponse)
def admin_dashboard(
    request: Request,
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin)
):
        
    stats = admin_service.get_admin_dashboard_stats(db)
    
    now = datetime.now()
    show_year_end_notice = now.month in (12, 1)
    next_year = now.year + 1 if now.month == 12 else now.year
    
    # 이미 내년도(혹은 새해) 연차 지급 설정이 데이터베이스에 존재한다면 배너를 자동으로 숨김 처리합니다.
    if show_year_end_notice:
        has_next_year_allocations = db.query(models.UserYearlyLeaveAllocations).filter(
            models.UserYearlyLeaveAllocations.year == next_year
        ).first() is not None
        if has_next_year_allocations:
            show_year_end_notice = False
    
    return _templates(request).TemplateResponse(request=request, name="admin_dashboard.html", context={
        "admin": admin,
        "active_users_count": stats["active_users_count"],
        "leaves_today_count": stats["leaves_today_count"],
        "pending_leaves_count": stats["pending_leaves_count"],
        "is_approval_required": stats["is_approval_required"],
        "today_used_hours": stats["today_used_hours"],
        "recent_leaves": stats["recent_leaves"],
        "current_year": now.year,
        "show_year_end_notice": show_year_end_notice,
        "next_year": next_year
    })

@page_router.get("/leave/timeline", response_class=HTMLResponse)
def admin_leaves_timeline(
    request: Request,
    year: int = None,
    month: int = 0,
    user_id: str = "",
    company: str = "",
    team: str = "",
    leave_status: str = "",
    sort: str = "created_at",
    sort_dir: str = "desc",
    order: str | None = None,
    page: int = 1,
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    current_year = year if year else datetime.now().year
    current_month = month
    selected_user_id = user_id.strip()
    selected_company = company.strip()
    selected_team = team.strip()
    selected_leave_status = _timeline_leave_status_filter(leave_status)

    sort_key = sort if sort in TIMELINE_SORT_COLUMNS else "created_at"
    if "sort" not in request.query_params and order in ("asc", "desc"):
        sort_dir = order
    sort_dir_effective = sort_dir if sort_dir in ("asc", "desc") else "desc"

    leave_year_rows = db.query(models.Leaves.year).distinct().all()
    leave_years = [row[0] for row in leave_year_rows]
    year_options = utils.build_year_options(datetime.now().year, leave_years)

    query = admin_service.get_leaves_timeline_query(
        db=db,
        year=current_year,
        month=current_month,
        user_id=selected_user_id,
        company=selected_company,
        team=selected_team,
        leave_status=selected_leave_status,
    )

    per_page = 50
    total_count = query.count()
    total_pages = (total_count + per_page - 1) // per_page
    if page > total_pages and total_pages > 0:
        page = total_pages
    if page < 1:
        page = 1

    sort_col = {
        "created_at": models.Leaves.created_at,
        "user_name": models.Users.user_name,
        "date": models.Leaves.date,
        "slot": models.Leaves.snapshot_slot_label,
        "company": models.Users.company,
        "team": models.Users.team,
    }[sort_key]

    if sort_dir_effective == "asc":
        query = query.order_by(sort_col.asc().nulls_last(), models.Leaves.id.asc())
    else:
        query = query.order_by(sort_col.desc().nulls_last(), models.Leaves.id.desc())

    leaves = query.offset((page - 1) * per_page).limit(per_page).all()

    all_users = (
        db.query(models.Users)
        .filter(models.Users.role != "ADMIN")
        .all()
    )
    users = sorted(all_users, key=lambda u: u.user_name.lower())
    company_options = sorted(list({u.company for u in all_users if u.company}))
    team_options = sorted(list({u.team for u in all_users if u.team}))

    path = request.url.path
    base_q = {"year": str(current_year), "month": str(current_month)}
    if selected_user_id:
        base_q["user_id"] = selected_user_id
    if selected_company:
        base_q["company"] = selected_company
    if selected_team:
        base_q["team"] = selected_team
    if selected_leave_status:
        base_q["leave_status"] = selected_leave_status

    sort_urls = {}
    for col in TIMELINE_SORT_COLUMNS:
        nxt = _timeline_next_sort_dir(col, sort_key, sort_dir_effective)
        q = {**base_q, "sort": col, "sort_dir": nxt}
        sort_urls[col] = f"{path}?{urlencode(q)}"

    chip_q_base = {"year": str(current_year), "month": str(current_month)}
    if selected_user_id:
        chip_q_base["user_id"] = selected_user_id
    if selected_company:
        chip_q_base["company"] = selected_company
    if selected_team:
        chip_q_base["team"] = selected_team

    def _timeline_chip_url(extra: dict) -> str:
        q = {
            **chip_q_base,
            "sort": sort_key,
            "sort_dir": sort_dir_effective,
            **extra,
        }
        q = {k: v for k, v in q.items() if v not in ("", None)}
        return f"{path}?{urlencode(q)}"

    timeline_chip_url_all = _timeline_chip_url({})
    timeline_chip_url_pending = _timeline_chip_url({"leave_status": "PENDING"})
    timeline_chip_url_approved = _timeline_chip_url({"leave_status": "APPROVED"})
    timeline_chip_url_rejected = _timeline_chip_url({"leave_status": "REJECTED"})

    return _templates(request).TemplateResponse(
        request=request,
        name="admin_leaves_timeline.html",
        context={
            "admin": admin,
            "leaves": leaves,
            "users": users,
            "company_options": company_options,
            "team_options": team_options,
            "selected_year": current_year,
            "selected_month": current_month,
            "selected_user_id": selected_user_id,
            "selected_company": selected_company,
            "selected_team": selected_team,
            "sort_key": sort_key,
            "sort_dir": sort_dir_effective,
            "sort_urls": sort_urls,
            "page": page,
            "total_pages": total_pages,
            "total_count": total_count,
            "export_query": urlencode(
                {
                    k: v
                    for k, v in {
                        "year": str(current_year),
                        "month": str(current_month),
                        "user_id": selected_user_id,
                        "company": selected_company,
                        "team": selected_team,
                        "leave_status": selected_leave_status,
                        "sort": sort_key,
                        "sort_dir": sort_dir_effective,
                    }.items()
                    if v not in ("", None)
                }
            ),
            "selected_leave_status": selected_leave_status,
            "timeline_chip_url_all": timeline_chip_url_all,
            "timeline_chip_url_pending": timeline_chip_url_pending,
            "timeline_chip_url_approved": timeline_chip_url_approved,
            "timeline_chip_url_rejected": timeline_chip_url_rejected,
            "current_year": datetime.now().year,
            "year_options": year_options,
        },
    )


@page_router.get("/leave/timeline/export")
def export_admin_leaves_timeline(
    request: Request,
    year: int = None,
    month: int = 0,
    user_id: str = "",
    company: str = "",
    team: str = "",
    leave_status: str = "",
    sort: str = "created_at",
    sort_dir: str = "desc",
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    current_year = year if year else datetime.now().year
    selected_user_id = user_id.strip()
    selected_company = company.strip()
    selected_team = team.strip()
    selected_leave_status = _timeline_leave_status_filter(leave_status)
    sort_key = sort if sort in TIMELINE_SORT_COLUMNS else "created_at"
    sort_dir_effective = sort_dir if sort_dir in ("asc", "desc") else "desc"

    query = (
        db.query(models.Leaves)
        .join(models.Users, models.Leaves.user_id == models.Users.user_id)
        .options(contains_eager(models.Leaves.user))
        .filter(models.Leaves.year == current_year)
    )
    if month > 0:
        import calendar as cal_module
        num_days = cal_module.monthrange(current_year, month)[1]
        query = query.filter(
            models.Leaves.date >= date_cls(current_year, month, 1),
            models.Leaves.date <= date_cls(current_year, month, num_days)
        )
    if selected_user_id:
        query = query.filter(models.Leaves.user_id == selected_user_id)
    if selected_company:
        query = query.filter(models.Users.company == selected_company)
    if selected_team:
        query = query.filter(models.Users.team == selected_team)
    if selected_leave_status:
        query = query.filter(models.Leaves.status == selected_leave_status)

    sort_col = {
        "created_at": models.Leaves.created_at,
        "user_name": models.Users.user_name,
        "date": models.Leaves.date,
        "slot": models.Leaves.snapshot_slot_label,
        "company": models.Users.company,
        "team": models.Users.team,
    }[sort_key]
    if sort_dir_effective == "asc":
        query = query.order_by(sort_col.asc().nulls_last(), models.Leaves.id.asc())
    else:
        query = query.order_by(sort_col.desc().nulls_last(), models.Leaves.id.desc())
    leaves = query.all()

    # 엑셀 다중 시트 작성 (WriteOnlyWorkbook 메모리 최적화 모드)
    wb = Workbook(write_only=True)
    
    # 1. 사원 리스트 추출 (필터 조건에 부합하는 사원들 대상)
    user_query = db.query(models.Users)
    if selected_user_id:
        user_query = user_query.filter(models.Users.user_id == selected_user_id)
    if selected_company:
        user_query = user_query.filter(models.Users.company == selected_company)
    if selected_team:
        user_query = user_query.filter(models.Users.team == selected_team)
    
    users = user_query.order_by(models.Users.company.asc(), models.Users.team.asc(), models.Users.user_name.asc()).all()
    user_ids = [u.user_id for u in users]

    # N+1 쿼리 방지: 배정량 Bulk 조회
    allocations = (
        db.query(models.UserYearlyLeaveAllocations)
        .filter(
            models.UserYearlyLeaveAllocations.user_id.in_(user_ids),
            models.UserYearlyLeaveAllocations.year == current_year
        )
        .all()
    )
    alloc_map = {a.user_id: a.allocated_hours for a in allocations}

    # N+1 쿼리 방지: 연도 차감 연차(is_deductive=True) 상태별 집계 조회
    leaves_summary = (
        db.query(
            models.Leaves.user_id,
            models.Leaves.status,
            func.sum(models.Leaves.snapshot_deduction_hours).label("total_hours")
        )
        .filter(
            models.Leaves.user_id.in_(user_ids),
            models.Leaves.year == current_year,
            models.Leaves.is_deductive == True
        )
        .group_by(models.Leaves.user_id, models.Leaves.status)
        .all()
    )

    leave_hours_map = {uid: {"APPROVED": 0.0, "PENDING": 0.0} for uid in user_ids}
    for uid, status, total_hours in leaves_summary:
        if uid in leave_hours_map and status in ("APPROVED", "PENDING"):
            leave_hours_map[uid][status] = float(total_hours or 0.0)

    # Sheet 1: 연차 현황 요약 작성
    ws_summary = wb.create_sheet(title="연차 현황 요약")
    ws_summary.append([

        "사원 ID", "사원명", "회사", "팀", "직급", "역할", "계정 상태",
        f"{current_year}년 총 배정 시간 (h)", "사용 완료 시간 (h)", "결재 대기 시간 (h)", "잔여 시간 (h)"
    ])

    for u in users:
        allocated = alloc_map.get(u.user_id, u.total_leave_hours or 120)
        approved = leave_hours_map.get(u.user_id, {}).get("APPROVED", 0.0)
        pending = leave_hours_map.get(u.user_id, {}).get("PENDING", 0.0)
        remaining = float(allocated) - approved
        
        ws_summary.append([
            u.user_id,
            u.user_name,
            u.company or "",
            u.team or "",
            u.position or "",
            u.role or "",
            "활성" if u.is_active else "비활성",
            float(allocated),
            float(approved),
            float(pending),
            float(remaining)
        ])

    # Sheet 2: 상세 신청 내역
    ws_timeline = wb.create_sheet(title="상세 신청 내역")
    ws_timeline.append([
        "신청 시각", "사용자명", "사용자 ID", "회사", "팀", "연차일", 
        "사용 시간대", "차감 여부", "차감시간 (h)", "상태", "신청 사유", "반려 사유"
    ])

    for leave in leaves:
        row = []
        
        # 1. 신청 시각 (datetime)
        if leave.created_at:
            cell_created = WriteOnlyCell(ws_timeline, value=utils.to_kst_naive(leave.created_at))
            cell_created.number_format = 'yyyy-mm-dd hh:mm:ss'
            row.append(cell_created)
        else:
            row.append("")
            
        # 2. 사용자명
        row.append(leave.user.user_name if leave.user else "")
        # 3. 사용자 ID
        row.append(leave.user.user_id if leave.user else "")
        # 4. 회사
        row.append((leave.user.company or "") if leave.user else "")
        # 5. 팀
        row.append((leave.user.team or "") if leave.user else "")
        
        # 6. 연차일 (date)
        if leave.date:
            cell_date = WriteOnlyCell(ws_timeline, value=leave.date)
            cell_date.number_format = 'yyyy-mm-dd'
            row.append(cell_date)
        else:
            row.append("")
            
        # 7. 사용 시간대
        row.append(leave.snapshot_slot_label or "")
        # 8. 차감 여부
        row.append("차감" if leave.is_deductive else "비차감")
        # 9. 차감시간 (h) (float)
        row.append(float(leave.snapshot_deduction_hours or 0.0))
        # 10. 상태
        row.append(leave.status or "")
        # 11. 신청 사유
        row.append(leave.reason or "")
        # 12. 반려 사유
        row.append(leave.rejection_reason or "")
        
        ws_timeline.append(row)

    output = io.BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"leave_summary_{current_year}_{stamp}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@page_router.get("/leave/calendar", response_class=HTMLResponse)
def admin_leaves_calendar(
    request: Request,
    year: int = None,
    month: int = None,
    view: str = "month",
    user_id: str = "",
    company: str = "",
    team: str = "",
    active_state: str = "all",
    sort: str = "",
    sort_dir: str = "asc",
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    current_year = year if year else datetime.now().year
    current_month = month if month else datetime.now().month
    if view not in ("month", "month_grid", "year"):
        view = "month"
    is_year_view = view == "year"
    selected_user_id = user_id.strip()
    selected_company = company.strip()
    selected_team = team.strip()
    selected_active_state = active_state if active_state in ("all", "active", "inactive") else "all"
    sort_dir_eff = sort_dir if sort_dir in ("asc", "desc") else "asc"
    sort_param = (sort or "").strip()
    sort_key_effective = sort_param if sort_param in CALENDAR_SORT_COLUMNS else None

    leave_year_rows = db.query(models.Leaves.year).distinct().all()
    leave_years = [row[0] for row in leave_year_rows]
    year_options = utils.build_year_options(datetime.now().year, leave_years)

    all_users = db.query(models.Users).filter(models.Users.role != "ADMIN").all()
    
    company_options = sorted(list({u.company for u in all_users if u.company}))
    team_options = sorted(list({u.team for u in all_users if u.team}))
    
    users_all = [
        u for u in all_users
        if selected_active_state == "all"
        or (selected_active_state == "active" and u.is_active)
        or (selected_active_state == "inactive" and not u.is_active)
    ]
    users_all.sort(key=lambda u: u.user_name.lower())
    
    users = [
        u for u in all_users
        if (selected_active_state == "all" or (selected_active_state == "active" and u.is_active) or (selected_active_state == "inactive" and not u.is_active))
        and (not selected_company or u.company == selected_company)
        and (not selected_team or u.team == selected_team)
        and (not selected_user_id or u.user_id == selected_user_id)
    ]

    weekday_labels = ["월", "화", "수", "목", "금", "토", "일"]

    user_leaves_map = {u.user_id: {} for u in users}
    days = []
    day_weekday_map = {}
    weekend_days = []
    holiday_day_map = {}
    user_month_hours = {u.user_id: {m: 0.0 for m in range(1, 13)} for u in users}

    user_ids = [u.user_id for u in users]
    user_yearly_leaves_map = {uid: [] for uid in user_ids}

    if user_ids:
        year_leaves = (
            db.query(models.Leaves)
            .filter(
                models.Leaves.user_id.in_(user_ids),
                models.Leaves.year == current_year,
                models.Leaves.status.notin_(["CANCELED", "REJECTED"]),
                models.Leaves.is_deductive == True
            )
            .all()
        )
        for l in year_leaves:
            user_yearly_leaves_map[l.user_id].append(l)

    offset = 0
    if is_year_view:
        for uid, leaves in user_yearly_leaves_map.items():
            for l in leaves:
                user_month_hours[uid][l.date.month] += float(l.snapshot_deduction_hours or 0)
    else:
        num_days = calendar.monthrange(current_year, current_month)[1]
        first_weekday_monday0 = calendar.weekday(current_year, current_month, 1)
        offset = (first_weekday_monday0 + 1) % 7
        days = list(range(1, num_days + 1))
        day_weekday_map = {
            d: weekday_labels[calendar.weekday(current_year, current_month, d)] for d in days
        }
        weekend_days = [d for d in days if calendar.weekday(current_year, current_month, d) >= 5]

        month_start = date_cls(current_year, current_month, 1)
        month_end = date_cls(current_year, current_month, num_days)
        month_holidays = (
            db.query(models.Holidays)
            .filter(models.Holidays.date >= month_start, models.Holidays.date <= month_end)
            .order_by(models.Holidays.date.asc())
            .all()
        )
        holiday_day_map = {h.date.day: h for h in month_holidays}

        month_leaves = []
        if user_ids:
            month_leaves = (
                db.query(models.Leaves)
                .filter(
                    models.Leaves.user_id.in_(user_ids),
                    models.Leaves.date >= month_start,
                    models.Leaves.date <= month_end,
                )
                .all()
            )
        for l in month_leaves:
            if l.user_id not in user_leaves_map:
                continue
            day = l.date.day
            if day not in user_leaves_map[l.user_id]:
                user_leaves_map[l.user_id][day] = []
            user_leaves_map[l.user_id][day].append(l)

    allocation_map = admin_service.get_yearly_allocation_map(db, [u.user_id for u in users], current_year)
    user_stats = []
    for u in users:
        if is_year_view:
            yearly_used = sum(user_month_hours[u.user_id].values())
            month_used_labels = {
                m: (
                    utils.hours_to_days_hours_label(user_month_hours[u.user_id][m])
                    if user_month_hours[u.user_id][m]
                    else "-"
                )
                for m in range(1, 13)
            }
            month_used_short_labels = {
                m: (
                    "-"
                    if not user_month_hours[u.user_id][m]
                    else utils.hours_to_days_hours_compact(user_month_hours[u.user_id][m])
                )
                for m in range(1, 13)
            }
        else:
            month_used_labels = {}
            month_used_short_labels = {}
            yearly_leaves = user_yearly_leaves_map.get(u.user_id, [])
            yearly_used = sum(float(l.snapshot_deduction_hours or 0) for l in yearly_leaves)

        allocated_hours = float(allocation_map.get(u.user_id, u.total_leave_hours or 0))
        yearly_remain = allocated_hours - float(yearly_used)

        if is_year_view:
            period_used = yearly_used
        else:
            period_used = 0.0
            if u.user_id in user_leaves_map:
                for dlist in user_leaves_map[u.user_id].values():
                    period_used += sum(float(lv.snapshot_deduction_hours or 0) for lv in dlist if lv.status not in ("CANCELED", "REJECTED") and lv.is_deductive == True)

        user_stats.append(
            {
                "user_id": u.user_id,
                "user_name": u.user_name,
                "company": u.company,
                "team": u.team,
                "total_hours": allocated_hours,
                "period_used": period_used,
                "yearly_used": yearly_used,
                "yearly_remain": yearly_remain,
                "yearly_remain_label": utils.hours_to_days_hours_label(yearly_remain),
                "period_used_label": utils.hours_to_days_hours_label(period_used),
                "yearly_remain_short": utils.hours_to_days_hours_compact(yearly_remain),
                "period_used_short": utils.hours_to_days_hours_compact(period_used),
                "month_used_labels": month_used_labels,
                "month_used_short_labels": month_used_short_labels,
                "is_active": u.is_active,
            }
        )

    if sort_key_effective is None:
        user_stats.sort(key=lambda x: (-int(x["is_active"]), x["user_name"].lower(), x["user_id"]))
    else:
        rev = sort_dir_eff == "desc"
        if sort_key_effective == "user_name":
            user_stats.sort(
                key=lambda x: (x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )
        elif sort_key_effective == "company":
            user_stats.sort(
                key=lambda x: ((x["company"] or "").lower(), x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )
        elif sort_key_effective == "team":
            user_stats.sort(
                key=lambda x: ((x["team"] or "").lower(), x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )
        elif sort_key_effective == "yearly_remain":
            user_stats.sort(
                key=lambda x: (x["yearly_remain"], x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )
        elif sort_key_effective == "period_used":
            user_stats.sort(
                key=lambda x: (x["period_used"], x["user_name"].lower(), x["user_id"]),
                reverse=rev,
            )

    path = request.url.path
    view_q = view
    base_q = {
        "year": str(current_year),
        "month": str(current_month),
        "view": view_q,
    }
    if selected_user_id:
        base_q["user_id"] = selected_user_id
    if selected_company:
        base_q["company"] = selected_company
    if selected_team:
        base_q["team"] = selected_team
    if selected_active_state != "all":
        base_q["active_state"] = selected_active_state

    sort_urls = {}
    for col in CALENDAR_SORT_COLUMNS:
        nxt = _calendar_next_sort_dir(col, sort_key_effective, sort_dir_eff)
        q = {**base_q, "sort": col, "sort_dir": nxt}
        sort_urls[col] = f"{path}?{urlencode(q)}"

    return _templates(request).TemplateResponse(
        request=request,
        name="admin_leaves_calendar.html",
        context={
            "admin": admin,
            "users": users_all,
            "user_leaves_map": dict(user_leaves_map),
            "user_stats": user_stats,
            "days": days,
            "day_weekday_map": day_weekday_map,
            "weekend_days": weekend_days,
            "holiday_day_map": holiday_day_map,
            "is_year_view": is_year_view,
            "view": view,
            "offset": offset,
            "selected_year": current_year,
            "selected_month": current_month,
            "selected_user_id": selected_user_id,
            "selected_company": selected_company,
            "selected_team": selected_team,
            "selected_active_state": selected_active_state,
            "company_options": company_options,
            "team_options": team_options,
            "sort_key": sort_key_effective,
            "sort_dir": sort_dir_eff,
            "sort_urls": sort_urls,
            "current_year": datetime.now().year,
            "year_options": year_options,
        },
    )

@page_router.get("/holidays", response_class=HTMLResponse)
def admin_holidays(request: Request, year: int = None, db: Session = Depends(get_db), admin: models.Users = Depends(get_current_admin)):

    now = datetime.now()
    current_year = year if year else now.year

    month_start = date_cls(current_year, 1, 1)
    month_end = date_cls(current_year, 12, 31)
    holidays = db.query(models.Holidays).filter(
        models.Holidays.date >= month_start,
        models.Holidays.date <= month_end
    ).order_by(models.Holidays.date.asc()).all()
    holiday_year_rows = db.query(extract('year', models.Holidays.date)).distinct().all()
    holiday_years = [int(row[0]) for row in holiday_year_rows if row[0] is not None]
    year_options = utils.build_year_options(now.year, holiday_years)

    return _templates(request).TemplateResponse(request=request, name="admin_holidays.html", context={
        "admin": admin,
        "holidays": holidays,
        "selected_year": current_year,
        "current_year": now.year,
        "year_options": year_options
    })

@api_router.post("/holiday/create")
def create_holiday(
    request: Request,
    holiday_name: str = Form(...),
    holiday_date: str = Form(...),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    holiday_name = holiday_name.strip()
    if not holiday_name:
        return JSONResponse(status_code=400, content={"message": "공휴일 이름을 입력해 주세요."})

    try:
        parsed_date = datetime.strptime(holiday_date, "%Y-%m-%d").date()
    except ValueError:
        return JSONResponse(status_code=400, content={"message": "날짜 형식이 올바르지 않습니다."})

    exist = db.query(models.Holidays).filter(models.Holidays.date == parsed_date).first()
    if exist:
        return JSONResponse(status_code=400, content={"message": "해당 날짜에는 이미 공휴일이 등록되어 있습니다."})

    new_holiday = models.Holidays(name=holiday_name, date=parsed_date)
    db.add(new_holiday)

    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="CREATE_HOLIDAY",
        target_info=f"Holiday:{parsed_date}",
        old_data="None",
        new_data=holiday_name
    )
    db.add(audit)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})

    return JSONResponse(status_code=200, content={"message": "공휴일이 등록되었습니다."})

@api_router.post("/holiday/update")
def update_holiday(
    request: Request,
    holiday_id: int = Form(...),
    holiday_name: str = Form(...),
    holiday_date: str = Form(...),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    holiday = db.query(models.Holidays).filter(models.Holidays.id == holiday_id).first()
    if not holiday:
        return JSONResponse(status_code=404, content={"message": "공휴일 정보를 찾을 수 없습니다."})

    holiday_name = holiday_name.strip()
    if not holiday_name:
        return JSONResponse(status_code=400, content={"message": "공휴일 이름을 입력해 주세요."})

    try:
        parsed_date = datetime.strptime(holiday_date, "%Y-%m-%d").date()
    except ValueError:
        return JSONResponse(status_code=400, content={"message": "날짜 형식이 올바르지 않습니다."})

    duplicate = db.query(models.Holidays).filter(
        models.Holidays.date == parsed_date,
        models.Holidays.id != holiday_id
    ).first()
    if duplicate:
        return JSONResponse(status_code=400, content={"message": "해당 날짜에는 이미 다른 공휴일이 등록되어 있습니다."})

    old_data = f"{holiday.date}:{holiday.name}"
    holiday.name = holiday_name
    holiday.date = parsed_date

    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="UPDATE_HOLIDAY",
        target_info=f"Holiday:{holiday_id}",
        old_data=old_data,
        new_data=f"{parsed_date}:{holiday_name}"
    )
    db.add(audit)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})

    return JSONResponse(status_code=200, content={"message": "공휴일이 수정되었습니다."})

@api_router.post("/holiday/delete")
def delete_holiday(
    request: Request,
    holiday_id: int = Form(...),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    holiday = db.query(models.Holidays).filter(models.Holidays.id == holiday_id).first()
    if not holiday:
        return JSONResponse(status_code=404, content={"message": "공휴일 정보를 찾을 수 없습니다."})

    old_data = f"{holiday.date}:{holiday.name}"
    db.delete(holiday)

    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="DELETE_HOLIDAY",
        target_info=f"Holiday:{holiday_id}",
        old_data=old_data,
        new_data="DELETED"
    )
    db.add(audit)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})

    return JSONResponse(status_code=200, content={"message": "공휴일이 삭제되었습니다."})

@api_router.post("/change-password")
def admin_change_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    if not auth.verify_password(current_password, admin.password):
        return JSONResponse(status_code=400, content={"message": "현재 비밀번호가 일치하지 않습니다."})

    if len(new_password) < 4:
        return JSONResponse(status_code=400, content={"message": "새 비밀번호는 최소 4자 이상이어야 합니다."})

    admin.password = auth.get_password_hash(new_password)
    admin.token_version += 1
    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="CHANGE_ADMIN_PASSWORD",
        target_info=f"Admin:{admin.user_id}",
        old_data="*****",
        new_data="*****"
    )
    db.add(audit)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})

    return JSONResponse(status_code=200, content={"message": "비밀번호가 성공적으로 변경되었습니다."})


@page_router.get("/users", response_class=HTMLResponse)
def admin_users(
    request: Request,
    filter: str = "all",
    year: int = None,
    sort_key: str = "role",
    sort_dir: str = "asc",
    q: str = "",
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
        
    query = db.query(models.Users)
    if filter == "active":
        query = query.filter(models.Users.is_active == True)
    elif filter == "inactive":
        query = query.filter(models.Users.is_active == False)
        
    users = query.all()

    # 인메모리 검색 적용 (Fernet 복호화 및 한글 초성 매칭)
    if q and q.strip():
        users = utils.search_users_stateless(users, q)

    sort_key_effective = sort_key if sort_key in {"user_name", "user_id", "company", "team", "leave_days", "role"} else "role"
    sort_dir_effective = "desc" if sort_dir == "desc" else "asc"
    now_year = datetime.now().year
    selected_year = year if year else now_year
    leave_year_rows = db.query(models.Leaves.year).distinct().all()
    leave_years = [row[0] for row in leave_year_rows]
    allocation_year_rows = db.query(models.UserYearlyLeaveAllocations.year).distinct().all()
    allocation_years = [row[0] for row in allocation_year_rows]
    year_options = utils.build_year_options(now_year, leave_years + allocation_years)
    allocation_map = admin_service.get_yearly_allocation_map(db, [u.user_id for u in users], selected_year)
    user_leave_days_map = {
        u.user_id: int(allocation_map.get(u.user_id, u.total_leave_hours or 0)) // 8
        for u in users
    }
    def _user_sort_tuple(u: models.Users):
        company_text = (u.company or "").lower()
        team_text = (u.team or "").lower()
        leave_days = user_leave_days_map.get(u.user_id, 0)
        role_priority = {"ADMIN": 0, "PM": 1, "TEAM_LEAD": 2, "STAFF": 3}
        key_map = {
            "user_name": ((u.user_name or "").lower(), (u.user_id or "").lower()),
            "user_id": ((u.user_id or "").lower(), (u.user_name or "").lower()),
            "company": (company_text, (u.user_name or "").lower(), (u.user_id or "").lower()),
            "team": (team_text, (u.user_name or "").lower(), (u.user_id or "").lower()),
            "leave_days": (leave_days, (u.user_name or "").lower(), (u.user_id or "").lower()),
            "role": (role_priority.get(u.role, 9), (u.user_name or "").lower(), (u.user_id or "").lower()),
        }
        return key_map[sort_key_effective]

    active_users = [u for u in users if u.is_active]
    inactive_users = [u for u in users if not u.is_active]

    active_sorted = sorted(active_users, key=_user_sort_tuple, reverse=(sort_dir_effective == "desc"))
    inactive_sorted = sorted(inactive_users, key=_user_sort_tuple, reverse=(sort_dir_effective == "desc"))
    users_sorted = active_sorted + inactive_sorted

    base_q = {"filter": filter, "year": str(selected_year)}
    if q:
        base_q["q"] = q
    sort_urls = {}
    for col in ("user_name", "user_id", "company", "team", "leave_days", "role"):
        next_dir = "desc" if (sort_key_effective == col and sort_dir_effective == "asc") else "asc"
        sort_params = {**base_q, "sort_key": col, "sort_dir": next_dir}
        sort_urls[col] = f"/admin/users?{urlencode(sort_params)}"
    
    return _templates(request).TemplateResponse(request=request, name="admin_users.html", context={
        "admin": admin,
        "users": users_sorted,
        "current_filter": filter,
        "sort_key": sort_key_effective,
        "sort_dir": sort_dir_effective,
        "sort_urls": sort_urls,
        "selected_year": selected_year,
        "year_options": year_options,
        "user_leave_days_map": user_leave_days_map,
        "q": q
    })

@api_router.post("/user/toggle")
def toggle_user_active(
    request: Request,
    background_tasks: BackgroundTasks,
    target_user_id: str = Form(...),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
        
    user = db.query(models.Users).filter(models.Users.user_id == target_user_id).first()
    if not user:
        return JSONResponse(status_code=404, content={"message": "User not found"})
        
    if user.user_id == admin.user_id:
        return JSONResponse(status_code=400, content={"message": "본인 계정은 비활성화 할 수 없습니다."})

    if user.role == "ADMIN":
        return JSONResponse(status_code=400, content={"message": "관리자 계정은 비활성화 할 수 없습니다."})
        
    old_status = user.is_active
    user.is_active = not user.is_active
    user.token_version += 1
    
    if not user.is_active:
        today = date_cls.today()
        future_leaves = db.query(models.Leaves).filter(
            models.Leaves.user_id == target_user_id,
            models.Leaves.date >= today,
            models.Leaves.status.notin_(["CANCELED", "REJECTED"])
        ).all()
        for leave in future_leaves:
            leave.status = "CANCELED"
            
    # Audit log
    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="TOGGLE_USER_ACTIVE",
        target_info=f"User:{target_user_id}",
        old_data=str(old_status),
        new_data=str(user.is_active)
    )
    db.add(audit)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    background_tasks.add_task(run_backup_and_rotate, DB_PATH)
    msg = "사원이 비활성화되었으며, 오늘 포함 미래 연차 신청 건이 일괄 취소되었습니다." if not user.is_active else "사원이 활성화되었습니다."
    return JSONResponse(status_code=200, content={"message": msg})

@api_router.post("/user/reset-password")
def reset_password(
    request: Request,
    background_tasks: BackgroundTasks,
    target_user_id: str = Form(...),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
        
    user = db.query(models.Users).filter(models.Users.user_id == target_user_id).first()
    if not user:
        return JSONResponse(status_code=404, content={"message": "User not found"})
        
    user.password = auth.get_password_hash("0000")
    user.token_version += 1
    
    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="RESET_PASSWORD",
        target_info=f"User:{target_user_id}",
        old_data="*****",
        new_data="0000"
    )
    db.add(audit)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    background_tasks.add_task(run_backup_and_rotate, DB_PATH)
    return JSONResponse(status_code=200, content={"message": "비밀번호가 '0000'으로 초기화되었습니다."})

@api_router.post("/leave/delete")
def delete_leave(
    request: Request,
    leave_id: int = Form(...),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
        
    leave = db.query(models.Leaves).filter(models.Leaves.id == leave_id).first()
    if not leave:
        return JSONResponse(status_code=404, content={"message": "Leave not found"})
        
    # Audit log (먼저 데이터 확보)
    actor_name = leave.user.user_name if leave.user else "알수없음"
    l_date = str(leave.date)
    l_status = str(leave.status)
    l_hours = str(leave.snapshot_deduction_hours)
    
    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="DELETE_LEAVE",
        target_info=f"Leave:{leave_id} ({actor_name}, {l_date})",
        old_data=f"Status:{l_status}, Date:{l_date}, Hours:{l_hours}",
        new_data="DELETED"
    )
    db.add(audit)
    db.delete(leave)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    background_tasks.add_task(run_backup_and_rotate, DB_PATH)
    return JSONResponse(status_code=200, content={"message": "연차가 성공적으로 삭제되었습니다."})


@api_router.post("/user/update")
def update_user(
    request: Request,
    target_user_id: str = Form(...),
    user_name: str = Form(...),
    company: str = Form(""),
    team: str = Form(""),
    role: str = Form("STAFF"),
    position: str = Form(""),
    is_active: bool = Form(True),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    user = db.query(models.Users).filter(models.Users.user_id == target_user_id).first()
    if not user:
        return JSONResponse(status_code=404, content={"message": "User not found"})

    if user.role == "ADMIN" and is_active is False:
        return JSONResponse(
            status_code=400, content={"message": "관리자 계정은 비활성화 할 수 없습니다."}
        )

    # role 검증
    from ..main import VALID_ROLES

    role_value = (role or "STAFF").strip().upper()
    if role_value not in VALID_ROLES:
        role_value = "STAFF"

    # ADMIN role은 직접 지정 불가
    if role_value == "ADMIN" and user.role != "ADMIN":
        return JSONResponse(
            status_code=400, content={"message": "관리자 역할은 직접 지정할 수 없습니다."}
        )

    old_data = (
        f"name={user.user_name};company={user.company};team={user.team};"
        f"role={user.role};position={user.position};active={user.is_active}"
    )

    user.user_name = user_name.strip()
    user.company = company.strip() if company else None
    user.team = team.strip() if team else None
    user.role = role_value
    user.position = position.strip() if position else None
    
    # 활성 상태 변경 감지 및 비활성화 시 미래 연차 일괄 취소
    was_active = user.is_active
    user.is_active = is_active
    user.token_version += 1

    if was_active and not user.is_active:
        today = date_cls.today()
        future_leaves = db.query(models.Leaves).filter(
            models.Leaves.user_id == target_user_id,
            models.Leaves.date >= today,
            models.Leaves.status.notin_(["CANCELED", "REJECTED"])
        ).all()
        for leave in future_leaves:
            leave.status = "CANCELED"

    new_data = (
        f"name={user.user_name};company={user.company};team={user.team};"
        f"role={user.role};position={user.position};active={user.is_active}"
    )

    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="UPDATE_USER_INFO",
        target_info=f"User:{target_user_id}",
        old_data=old_data,
        new_data=new_data,
    )
    db.add(audit)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    background_tasks.add_task(run_backup_and_rotate, DB_PATH)
    msg = "사원 정보가 수정되었으며, 비활성화 처리로 인해 오늘 포함 미래 연차 신청 건이 일괄 취소되었습니다." if (was_active and not user.is_active) else "사원 정보가 성공적으로 수정되었습니다."
    return JSONResponse(status_code=200, content={"message": msg})


@api_router.post("/user/create")
def create_user(
    request: Request,
    user_id: str = Form(...),
    user_name: str = Form(...),
    company: str = Form(""),
    team: str = Form(""),
    role: str = Form("STAFF"),
    position: str = Form(""),
    total_leave_days: int = Form(15),
    year: int = Form(None),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
        
    user_id = user_id.strip()
    if total_leave_days < 0:
        return JSONResponse(status_code=400, content={"message": "연차일수는 0 이상이어야 합니다."})

    # role 검증
    from ..main import VALID_ROLES
    role_value = (role or "STAFF").strip().upper()
    if role_value not in VALID_ROLES:
        role_value = "STAFF"
    # ADMIN role은 사용자 생성에서 지정 불가
    if role_value == "ADMIN":
        return JSONResponse(status_code=400, content={"message": "관리자 역할은 직접 지정할 수 없습니다."})
    position_value = (position or "").strip()[:60]

    exist = db.query(models.Users).filter(models.Users.user_id == user_id).first()
    if exist:
        return JSONResponse(status_code=400, content={"message": "이미 존재하는 ID입니다."})
        
    target_year = year if year else datetime.now().year
    new_user = models.Users(
        user_id=user_id,
        user_name=user_name,
        company=company,
        team=team,
        total_leave_hours=total_leave_days * 8,
        password=auth.get_password_hash("0000"),
        is_active=True,
        role=role_value,
        position=position_value if position_value else None,
    )
    db.add(new_user)
    db.flush()
    db.add(
        models.UserYearlyLeaveAllocations(
            user_id=user_id,
            year=target_year,
            allocated_hours=total_leave_days * 8
        )
    )
    
    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="CREATE_USER",
        target_info=f"User:{user_id}",
        old_data="None",
        new_data=f"{user_name};role={role_value}"
    )
    db.add(audit)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    background_tasks.add_task(run_backup_and_rotate, DB_PATH)
    return JSONResponse(status_code=200, content={"message": f"{user_name} 등록 완료! 초기 비번: 0000"})

@api_router.post("/user/update-leave-days")
def update_user_leave_days(
    request: Request,
    target_user_id: str = Form(...),
    total_leave_days: int = Form(...),
    year: int = Form(None),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    if total_leave_days < 0:
        return JSONResponse(status_code=400, content={"message": "연차일수는 0 이상이어야 합니다."})

    user = db.query(models.Users).filter(models.Users.user_id == target_user_id).first()
    if not user:
        return JSONResponse(status_code=404, content={"message": "User not found"})

    target_year = year if year else datetime.now().year
    old_hours = user.total_leave_hours
    new_hours = total_leave_days * 8
    allocation = db.query(models.UserYearlyLeaveAllocations).filter(
        models.UserYearlyLeaveAllocations.user_id == target_user_id,
        models.UserYearlyLeaveAllocations.year == target_year
    ).first()
    old_allocated = int(allocation.allocated_hours) if allocation else int(old_hours or 0)
    if allocation:
        allocation.allocated_hours = new_hours
    else:
        db.add(
            models.UserYearlyLeaveAllocations(
                user_id=target_user_id,
                year=target_year,
                allocated_hours=new_hours
            )
        )

    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="UPDATE_USER_LEAVE_DAYS",
        target_info=f"User:{target_user_id}",
        old_data=f"{target_year}:{old_allocated}h",
        new_data=f"{target_year}:{new_hours}h({total_leave_days}d)"
    )
    db.add(audit)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    background_tasks.add_task(run_backup_and_rotate, DB_PATH)
    return JSONResponse(status_code=200, content={"message": "연차일수가 변경되었습니다."})


@api_router.post("/user/bulk-update-leave-days")
def bulk_update_user_leave_days(
    request: Request,
    total_leave_days: int = Form(...),
    year: int = Form(None),
    filter_scope: str = Form("all"),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    if total_leave_days < 0:
        return JSONResponse(status_code=400, content={"message": "연차일수는 0 이상이어야 합니다."})

    target_year = year if year else datetime.now().year
    new_hours = total_leave_days * 8

    q = db.query(models.Users).filter(models.Users.role != "ADMIN")
    if filter_scope == "active":
        q = q.filter(models.Users.is_active == True)
    elif filter_scope == "inactive":
        q = q.filter(models.Users.is_active == False)
    target_users = q.all()

    if not target_users:
        return JSONResponse(status_code=400, content={"message": "일괄지급 대상 사용자가 없습니다."})

    updated_count = 0
    for user in target_users:
        allocation = db.query(models.UserYearlyLeaveAllocations).filter(
            models.UserYearlyLeaveAllocations.user_id == user.user_id,
            models.UserYearlyLeaveAllocations.year == target_year
        ).first()
        if allocation:
            allocation.allocated_hours = new_hours
        else:
            db.add(
                models.UserYearlyLeaveAllocations(
                    user_id=user.user_id,
                    year=target_year,
                    allocated_hours=new_hours
                )
            )
        updated_count += 1

    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="BULK_UPDATE_USER_LEAVE_DAYS",
        target_info=f"Scope:{filter_scope}, Year:{target_year}, Count:{updated_count}",
        old_data="BULK",
        new_data=f"{new_hours}h({total_leave_days}d)"
    )
    db.add(audit)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    background_tasks.add_task(run_backup_and_rotate, DB_PATH)
    return JSONResponse(
        status_code=200,
        content={"message": f"{target_year}년 연차일수를 {updated_count}명에게 일괄 {total_leave_days}일로 반영했습니다."}
    )

@api_router.post("/user/hard-delete")
def hard_delete_user(
    request: Request,
    target_user_id: str = Form(...),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
        
    user = db.query(models.Users).filter(models.Users.user_id == target_user_id).first()
    if not user:
        return JSONResponse(status_code=404, content={"message": "User not found"})
        
    if user.user_id == admin.user_id:
        return JSONResponse(status_code=400, content={"message": "본인 계정은 완전히 삭제할 수 없습니다."})

    if user.role == "ADMIN":
        return JSONResponse(status_code=400, content={"message": "관리자 계정은 삭제할 수 없습니다."})
        
    # 삭제하려는 사용자의 모든 연차 내역 먼저 삭제 (외래키 제약조건 방지)
    db.query(models.Leaves).filter(models.Leaves.user_id == target_user_id).delete()
    db.query(models.UserYearlyLeaveAllocations).filter(
        models.UserYearlyLeaveAllocations.user_id == target_user_id
    ).delete()
    db.query(models.AuditLogs).filter(models.AuditLogs.actor_id == target_user_id).update(
        {models.AuditLogs.actor_id: None}, synchronize_session=False
    )
    
    user_name = user.user_name
    db.delete(user)
    
    # Audit log
    audit = models.AuditLogs(
        actor_id=admin.user_id,
        action="HARD_DELETE_USER",
        target_info=f"User:{target_user_id} ({user_name})",
        old_data="EXISTS",
        new_data="DELETED"
    )
    db.add(audit)
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    background_tasks.add_task(run_backup_and_rotate, DB_PATH)
    return JSONResponse(status_code=200, content={"message": "사원 계정이 완전히 삭제되었습니다."})


@api_router.post("/leave/update-status")
def update_leave_status(
    request: Request,
    leave_id: int = Form(...),
    status_value: str = Form(...),
    rejection_reason: str = Form(""),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    leave = db.query(models.Leaves).filter(models.Leaves.id == leave_id).first()
    if not leave:
        return JSONResponse(status_code=404, content={"message": "Leave not found"})

    try:
        transition = apply_leave_status_transition(
            leave=leave,
            status_value=status_value,
            rejection_reason=rejection_reason,
        )
    except LeaveStatusTransitionError as exc:
        return JSONResponse(status_code=400, content={"message": str(exc)})

    leave = db.query(models.Leaves).filter(models.Leaves.id == leave_id).first()
    db.add(
        models.AuditLogs(
            actor_id=admin.user_id,
            action="UPDATE_LEAVE_STATUS",
            target_info=f"Leave:{leave_id} ({leave.user.user_name if leave else 'Unknown'}, {leave.date if leave else 'Unknown'})",
            old_data=transition.audit_old_data,
            new_data=transition.audit_new_data,
        )
    )
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    background_tasks.add_task(run_backup_and_rotate, DB_PATH)
    return JSONResponse(status_code=200, content={"message": "상태가 변경되었습니다."})


@api_router.post("/leave/update-type")
def update_leave_type(
    request: Request,
    leave_id: int = Form(...),
    is_deductive: bool = Form(...),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    leave = db.query(models.Leaves).filter(models.Leaves.id == leave_id).first()
    if not leave:
        return JSONResponse(status_code=404, content={"message": "신청 건을 찾을 수 없습니다."})

    old_type = "연차(차감)" if leave.is_deductive else "공가/출장(비차감)"
    new_type = "연차(차감)" if is_deductive else "공가/출장(비차감)"

    if leave.is_deductive == is_deductive:
        return JSONResponse(status_code=400, content={"message": "이미 해당 유형입니다."})

    # 비차감 -> 차감 변경 시 잔여 연차 체크
    if is_deductive:
        total_allocated = resolve_user_yearly_allocated_hours(db, leave.user, leave.year)
        used_hours = db.query(func.sum(models.Leaves.snapshot_deduction_hours)).filter(
            models.Leaves.user_id == leave.user_id,
            models.Leaves.year == leave.year,
            models.Leaves.status.notin_(["CANCELED", "REJECTED"]),
            models.Leaves.is_deductive == True,
            models.Leaves.id != leave.id
        ).scalar() or 0.0
        
        if used_hours + leave.snapshot_deduction_hours > total_allocated:
            return JSONResponse(status_code=400, content={"message": "사용자의 잔여 연차가 부족하여 연차로 변경할 수 없습니다."})

    leave.is_deductive = is_deductive
    
    db.add(
        models.AuditLogs(
            actor_id=admin.user_id,
            action="UPDATE_LEAVE_TYPE",
            target_info=f"Leave:{leave_id} ({leave.user.user_name}, {leave.date})",
            old_data=f"is_deductive={not is_deductive} ({old_type})",
            new_data=f"is_deductive={is_deductive} ({new_type})",
        )
    )
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    background_tasks.add_task(run_backup_and_rotate, DB_PATH)
    return JSONResponse(status_code=200, content={"message": f"유형이 {new_type}로 변경되었습니다."})


@api_router.get("/settings/approval")
def get_approval_setting(request: Request, db: Session = Depends(get_db), admin: models.Users = Depends(get_current_admin)):
    setting = _ensure_system_setting(db)
    return JSONResponse(
        status_code=200,
        content={
            "is_approval_required": setting.is_approval_required,
            "time_granularity_minutes": setting.time_granularity_minutes,
            "work_start_minute": setting.work_start_minute,
            "work_end_minute": setting.work_end_minute,
            "lunch_start_minute": setting.lunch_start_minute,
            "lunch_end_minute": setting.lunch_end_minute,
            "product_display_name": getattr(setting, "product_display_name", None) or "SHIM",
            "product_nav_short": getattr(setting, "product_nav_short", None) or "",
            "brand_initial": getattr(setting, "brand_initial", None) or "S",
            "team_calendar_visible": bool(getattr(setting, "team_calendar_visible", True)),
            "company_calendar_visible": bool(getattr(setting, "company_calendar_visible", False)),
        },
    )


BRANDING_FIELD_MAX = 120
BRANDING_NAV_SHORT_MAX = 80
BRANDING_BADGE_MAX = 24


@api_router.post("/settings/branding")
def set_branding_setting(
    request: Request,
    product_display_name: str = Form(...),
    product_nav_short: str = Form(""),
    brand_initial: str = Form(""),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    display = (product_display_name or "").strip()
    nav_short = (product_nav_short or "").strip()
    initial_raw = (brand_initial or "").strip()[:BRANDING_BADGE_MAX]

    if not display or len(display) > BRANDING_FIELD_MAX:
        return JSONResponse(status_code=400, content={"message": "공식 프로젝트 이름은 1~120자로 입력해 주세요."})
    if len(nav_short) > BRANDING_NAV_SHORT_MAX:
        return JSONResponse(status_code=400, content={"message": f"상단 바 짧은 이름은 {BRANDING_NAV_SHORT_MAX}자 이하로 입력해 주세요."})

    initial = initial_raw
    if not initial:
        initial = (display[:1] if display else "L") or "L"

    setting = _ensure_system_setting(db)
    old_data = (
        f"display={setting.product_display_name};"
        f"nav_short={setting.product_nav_short};"
        f"initial={setting.brand_initial}"
    )
    setting.product_display_name = display
    setting.product_nav_short = nav_short
    setting.brand_initial = initial
    new_data = f"display={display};nav_short={nav_short};initial={initial}"

    db.add(
        models.AuditLogs(
            actor_id=admin.user_id,
            action="UPDATE_BRANDING_SETTING",
            target_info="SystemSettings",
            old_data=old_data,
            new_data=new_data,
        )
    )
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    get_system_settings(db, force_reload=True)
    return JSONResponse(status_code=200, content={"message": "브랜딩 설정이 저장되었습니다."})


@api_router.post("/settings/calendar-scope")
def set_calendar_scope_setting(
    request: Request,
    scope: str = Form(...),  # "none", "team", "company"
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    if scope not in ("none", "team", "company"):
        return JSONResponse(status_code=400, content={"detail": "Invalid scope value"})

    setting = _ensure_system_setting(db)

    old_team = bool(getattr(setting, "team_calendar_visible", True))
    old_company = bool(getattr(setting, "company_calendar_visible", False))
    if old_company:
        old_scope = "company"
    elif old_team:
        old_scope = "team"
    else:
        old_scope = "none"

    if scope == "none":
        setting.team_calendar_visible = False
        setting.company_calendar_visible = False
    elif scope == "team":
        setting.team_calendar_visible = True
        setting.company_calendar_visible = False
    elif scope == "company":
        setting.team_calendar_visible = True
        setting.company_calendar_visible = True

    db.add(
        models.AuditLogs(
            actor_id=admin.user_id,
            action="UPDATE_CALENDAR_SCOPE_SETTING",
            target_info="SystemSettings",
            old_data=old_scope,
            new_data=scope,
        )
    )
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    get_system_settings(db, force_reload=True)
    return JSONResponse(status_code=200, content={"message": "캘린더 공유 범위 설정이 변경되었습니다."})




@api_router.post("/settings/approval")
def set_approval_setting(
    request: Request,
    is_approval_required: bool = Form(...),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
    setting = _ensure_system_setting(db)
    setting.is_approval_required = is_approval_required
    db.add(
        models.AuditLogs(
            actor_id=admin.user_id,
            action="UPDATE_APPROVAL_SETTING",
            target_info="SystemSettings",
            old_data="",
            new_data=str(is_approval_required),
        )
    )
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    get_system_settings(db, force_reload=True)
    return JSONResponse(status_code=200, content={"message": "승인 설정이 변경되었습니다."})


@api_router.post("/settings/time-policy")
def set_time_policy(
    request: Request,
    time_granularity_minutes: int = Form(...),
    work_start_minute: int = Form(...),
    work_end_minute: int = Form(...),
    lunch_start_minute: int = Form(-1),
    lunch_end_minute: int = Form(-1),
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    if time_granularity_minutes not in ALLOWED_TIME_GRANULARITIES:
        return JSONResponse(status_code=400, content={"message": "시간 단위는 30/60/120분만 허용됩니다."})
    if work_start_minute < 0 or work_start_minute > 1439 or work_end_minute < 1 or work_end_minute > 1440:
        return JSONResponse(status_code=400, content={"message": "업무시간 값이 올바르지 않습니다."})
    if work_start_minute % 30 != 0 or work_end_minute % 30 != 0:
        return JSONResponse(status_code=400, content={"message": "업무시간은 30분 단위로 설정해 주세요."})
    if work_end_minute <= work_start_minute:
        return JSONResponse(status_code=400, content={"message": "업무 종료 시간은 시작 시간보다 늦어야 합니다."})

    normalized_lunch_start = None if lunch_start_minute < 0 else lunch_start_minute
    normalized_lunch_end = None if lunch_end_minute < 0 else lunch_end_minute
    if (normalized_lunch_start is None) != (normalized_lunch_end is None):
        return JSONResponse(status_code=400, content={"message": "점심시간 제외는 시작/종료를 함께 입력해야 합니다."})
    if normalized_lunch_start is not None:
        if normalized_lunch_start < 0 or normalized_lunch_start > 1439 or normalized_lunch_end < 0 or normalized_lunch_end > 1439:
            return JSONResponse(status_code=400, content={"message": "점심시간 설정 값이 올바르지 않습니다."})
        if normalized_lunch_end <= normalized_lunch_start:
            return JSONResponse(status_code=400, content={"message": "점심시간 종료는 시작보다 늦어야 합니다."})

    setting = _ensure_system_setting(db)
    old_data = (
        f"granularity={setting.time_granularity_minutes};"
        f"work={setting.work_start_minute}-{setting.work_end_minute};"
        f"lunch={setting.lunch_start_minute}-{setting.lunch_end_minute}"
    )
    setting.time_granularity_minutes = time_granularity_minutes
    setting.work_start_minute = work_start_minute
    setting.work_end_minute = work_end_minute
    setting.lunch_start_minute = normalized_lunch_start
    setting.lunch_end_minute = normalized_lunch_end
    new_data = (
        f"granularity={setting.time_granularity_minutes};"
        f"work={setting.work_start_minute}-{setting.work_end_minute};"
        f"lunch={setting.lunch_start_minute}-{setting.lunch_end_minute}"
    )

    db.add(
        models.AuditLogs(
            actor_id=admin.user_id,
            action="UPDATE_TIME_POLICY_SETTING",
            target_info="SystemSettings",
            old_data=old_data,
            new_data=new_data,
        )
    )
    try:
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        return JSONResponse(status_code=500, content={"message": utils.format_db_error_message(e)})
    get_system_settings(db, force_reload=True)
    return JSONResponse(status_code=200, content={"message": "시간 단위/점심시간 정책이 저장되었습니다."})


@api_router.post("/ops/backup")
def run_backup(request: Request, db: Session = Depends(get_db), admin: models.Users = Depends(get_current_admin)):

    backup_path = create_sqlite_backup(DB_PATH, DB_PATH.parent / "backup")
    db.add(
        models.AuditLogs(
            actor_id=admin.user_id,
            action="MANUAL_DB_BACKUP",
            target_info="Database",
            old_data="",
            new_data=str(backup_path.name),
        )
    )
    db.commit()
    return JSONResponse(status_code=200, content={"message": f"백업이 생성되었습니다: {backup_path.name}"})


@page_router.get("/master", response_class=HTMLResponse)
def admin_master(request: Request, db: Session = Depends(get_db), admin: models.Users = Depends(get_current_admin)):
    setting = _ensure_system_setting(db)
    is_approval_required = bool(setting.is_approval_required) if setting else False
    return _templates(request).TemplateResponse(
        request=request,
        name="admin_master.html",
        context={
            "admin": admin,
            "current_year": datetime.now().year,
            "is_approval_required": is_approval_required,
            "time_granularity_minutes": int(getattr(setting, "time_granularity_minutes", 60) or 60),
            "work_start_minute": int(getattr(setting, "work_start_minute", 9 * 60) or (9 * 60)),
            "work_end_minute": int(getattr(setting, "work_end_minute", 18 * 60) or (18 * 60)),
            "lunch_start_minute": getattr(setting, "lunch_start_minute", None),
            "lunch_end_minute": getattr(setting, "lunch_end_minute", None),
            "half_hour_options": utils.build_half_hour_options(),
            "team_calendar_visible": bool(getattr(setting, "team_calendar_visible", True)),
            "company_calendar_visible": bool(getattr(setting, "company_calendar_visible", False)),
        },
    )


AUDIT_ACTION_LABELS = {
    "CREATE_USER": "신규 사원 등록",
    "UPDATE_USER_INFO": "사원 정보 수정",
    "DELETE_USER": "사원 계정 삭제",
    "HARD_DELETE_USER": "사원 정보 영구 삭제",
    "RESET_PASSWORD": "사원 비밀번호 초기화",
    "TOGGLE_USER_ACTIVE": "사원 활성 상태 변경",
    "UPDATE_USER_LEAVE_DAYS": "연차 일수 수정",
    "BULK_UPDATE_USER_LEAVE_DAYS": "연차 일수 일괄 수정",
    "UPDATE_APPROVAL_SETTING": "승인 워크플로우 설정 변경",
    "UPDATE_TEAM_CALENDAR_SETTING": "팀 캘린더 공유 설정 변경",
    "UPDATE_COMPANY_CALENDAR_SETTING": "전사 캘린더 공유 설정 변경",
    "UPDATE_CALENDAR_SCOPE_SETTING": "캘린더 공유 범위 변경",
    "UPDATE_TIME_POLICY_SETTING": "시간 정책 변경",
    "UPDATE_BRANDING_SETTING": "브랜딩 설정 변경",
    "MANUAL_DB_BACKUP": "수동 DB 백업",
    "CREATE_HOLIDAY": "공휴일 등록",
    "UPDATE_HOLIDAY": "공휴일 정보 수정",
    "DELETE_HOLIDAY": "공휴일 삭제",
    "CHANGE_ADMIN_PASSWORD": "관리자 비밀번호 변경",
    "CHANGE_PASSWORD": "사용자 비밀번호 변경",
    "DELETE_LEAVE": "연차 신청 삭제",
    "UPDATE_LEAVE_STATUS": "결재 상태 변경",
    "APPROVE_LEAVE": "연차 승인(팀장/PM)",
    "REJECT_LEAVE": "연차 반려(팀장/PM)",
    "APPLY_LEAVE_BULK": "다수일 연차 일괄 신청",
    "UPDATE_LEAVE_TYPE": "휴가 유형 정정(연차↔출장/공가)",
}


def get_audit_action_label(action: str) -> str:
    if not action:
        return ""
    label = AUDIT_ACTION_LABELS.get(action)
    if not label and action.startswith("SEED_KR_HOLIDAYS_"):
        year_part = action.replace("SEED_KR_HOLIDAYS_", "")
        return f"{year_part}년 공휴일 자동 생성"
    return label or action


def get_audit_target_label(target_info: str) -> str:
    raw_target = (target_info or "").strip()
    if not raw_target:
        return ""

    target = raw_target
    if target.startswith("User:"):
        target = target.replace("User:", "사원:")
    elif target.startswith("Admin:"):
        target = target.replace("Admin:", "관리자:")
    elif target.startswith("Holiday:"):
        target = target.replace("Holiday:", "공휴일:")
    elif target.startswith("Leave:"):
        if "(" in target and ")" in target:
            parts = target.replace("Leave:", "").split(" (", 1)
            leave_id_part = parts[0]
            detail_part = parts[1].replace(")", "")
            target = f"연차신청:{detail_part} [ID:{leave_id_part}]"
        else:
            target = target.replace("Leave:", "연차신청(ID:") + ")"
    elif target.startswith("HolidaySeed:"):
        target = target.replace("HolidaySeed:", "") + "년 공휴일 생성"
    elif target == "SystemSettings":
        target = "시스템 설정"
    elif target == "Database":
        target = "데이터베이스"
    elif target.startswith("Scope:"):
        # Scope:all, Year:2026, Count:5
        parts = {}
        for part in target.split(","):
            if ":" in part:
                k, v = part.split(":", 1)
                parts[k.strip()] = v.strip()
        scope_val = parts.get("Scope", "all")
        scope_ko = {"all": "전체", "active": "활성 사원", "inactive": "비활성 사원"}.get(scope_val, scope_val)
        year_val = parts.get("Year", "")
        count_val = parts.get("Count", "")
        target = f"범위:{scope_ko}, 연도:{year_val}, 건수:{count_val}"
    elif target.startswith("Leaves for "):
        dates_str = target.replace("Leaves for ", "")
        target = f"다수일 연차 신청: {dates_str}"

    return target


@page_router.get("/audit", response_class=HTMLResponse)
def admin_audit_logs(
    request: Request,
    actor_id: str = "",
    action: str = "",
    start_date: str = "",
    end_date: str = "",
    page: int = 1,
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):

    per_page = 50
    s_date = None
    e_date = None
    if start_date:
        try:
            s_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            pass
    if end_date:
        try:
            e_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            pass

    query = admin_service.get_audit_logs_query(
        db=db,
        actor_id=actor_id.strip(),
        action=action.strip(),
        start_date=s_date,
        end_date=e_date,
    )

    total_count = query.count()
    total_pages = (total_count + per_page - 1) // per_page
    if page > total_pages and total_pages > 0:
        page = total_pages
    if page < 1:
        page = 1

    logs = (
        query.order_by(models.AuditLogs.timestamp.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    # 필터용 사용자 목록
    users = db.query(models.Users).filter(models.Users.is_active == True).order_by(models.Users.user_name).all()

    # 가독성 라벨 추가
    for log in logs:
        log.action_label = get_audit_action_label(log.action)
        log.target_label = get_audit_target_label(log.target_info) or (log.target_info or "")

    export_q = {
        "actor_id": actor_id,
        "action": action,
        "start_date": start_date,
        "end_date": end_date,
    }

    return _templates(request).TemplateResponse(
        request=request,
        name="admin_audit.html",
        context={
            "admin": admin,
            "logs": logs,
            "users": users,
            "actor_id": actor_id,
            "action": action,
            "start_date": start_date,
            "end_date": end_date,
            "page": page,
            "total_pages": total_pages,
            "total_count": total_count,
            "current_year": datetime.now().year,
            "export_query": urlencode({k: v for k, v in export_q.items() if v}),
        },
    )


@page_router.get("/audit/export")
def admin_audit_export(
    actor_id: str = "",
    action: str = "",
    start_date: str = "",
    end_date: str = "",
    db: Session = Depends(get_db),
    admin: models.Users = Depends(get_current_admin),
):
    s_date = None
    e_date = None
    if start_date:
        try:
            s_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            pass
    if end_date:
        try:
            e_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            pass

    query = admin_service.get_audit_logs_query(
        db=db,
        actor_id=actor_id.strip(),
        action=action.strip(),
        start_date=s_date,
        end_date=e_date,
    )
    logs = query.order_by(models.AuditLogs.timestamp.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"
    headers = ["시각", "수행자(ID)", "수행자(이름)", "수행자 소속", "액션(코드)", "액션(내용)", "대상 정보", "이전 데이터", "이후 데이터"]
    ws.append(headers)

    for log in logs:
        actor_name = log.actor.user_name if log.actor else (log.actor_name if log.actor_name else "System")
        if log.actor_department:
            actor_dept = log.actor_department
        elif log.actor:
            actor_dept = f"{log.actor.company or ''} {log.actor.team or ''}".strip()
        else:
            actor_dept = ""
        action_label = get_audit_action_label(log.action)
        target = get_audit_target_label(log.target_info) or (log.target_info or "")

        ws.append(
            [
                utils.format_datetime_kst(log.timestamp, "%Y-%m-%d %H:%M:%S"),
                log.actor_id or "",
                actor_name,
                actor_dept,
                log.action,
                action_label,
                target,
                log.old_data,
                log.new_data,
            ]
        )

    out = io.BytesIO()
    wb.save(out)
    wb.close()
    out.seek(0)

    filename = f"audit_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


